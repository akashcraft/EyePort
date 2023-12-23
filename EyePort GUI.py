#EyePort Source Code
#Made by Akash Samanta

import asyncio,time,_thread
from math import floor,ceil
from threading import Thread
import g3pylib
from queue import Empty

import tkinter as tk
from customtkinter import *
from tkinter import ttk
from tkinter import Label
from PIL import ImageTk, Image
import os,subprocess,cv2
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import openpyxl as xls
import csv
import AppOpener,webbrowser
from tkVideoPlayer import TkinterVideo

from Data_Extraction_Rotating_Head import extract
from Data_Analysis_Rotating_Head import analyzer,progress_queue
from ModelCreator import creator
from GIF import ImageLabel
import logging
imageai_avail = True
try:
    from imageai.Detection.Custom import CustomObjectDetection
    from imageai.Detection import ObjectDetection
except:
    imageai_avail = False

progress = 0
exp = True #Boolean for Picture Overlays
compact = 0 #Collapse Side Bar
logging.getLogger('matplotlib.font_manager').disabled = True
welcome = True
fullscreenmode = 0
root = CTk()
settings = []
recordactive = False
connected = False
disconnect = False
asense = 0.2
popenabled = True
exportsavefolder = ""
loaddata = ""
loadeddata = False
rememberloadbool = True
useai = 0
exportstatus = 3
overlay = 0
aviolate = 10
language_mode = 0
headsen = 20

def main():
    global framebuttons
    #Settings Creator and Default Settings
    def createsettings():
        firstsettings = "EyePort Settings File - Any corruption may lead to configuration loss\nMade by Akash Samanta\nLight\n100\n20\n30\n40\n0\n300\n0\n0\n0\nNormal\n0\n1\nLoad1\nLoad2\n200\n0\nExpsave\n2\n0\n0\n0\n0\n"
        try:
            f=open("Settings.txt","w")
        except PermissionError:
            tk.messagebox.showerror("EyePort Administrator Privileges Needed","Because of the nature of this EyePort install, you will need to launch EyePort as administrator. To prevent this behaviour, please install EyePort again on a user directory.")
            quit()
        f.write(firstsettings)
        f.close()

    #Record Page Functions
    #Battery Indicator
    async def batteryrefresh():
        l = await tobiiglasses.system.battery.get_level()
        l = float(str(l*100)[:4])
        text21.configure(text=f"{l}%")
        if l>95.0:
            batteryicon.configure(image=imgtk9)
        elif l>55.0:
            batteryicon.configure(image=imgtk10)
        elif l>35.0:
            batteryicon.configure(image=imgtk11)
        elif l>10.0:
            batteryicon.configure(image=imgtk12)
        else:
            batteryicon.configure(image=imgtk13)
        timelefticon.configure(image=imgtk14)
        h=1.8*l
        m=int(h%60)
        h=int(h/60)
        text22.configure(text=f"{h}h {m}m")
        text21.grid(row=0,column=2,pady=(10,0),padx=10,sticky='w')
        text22.grid(row=0,column=4,pady=(10,0),padx=10,sticky='w')
        batteryicon.grid(row=0,column=1,pady=(10,0),sticky='e')
        timelefticon.grid(row=0,column=3,pady=(10,0),sticky='e')

    #Connection to Glasses
    def connectsuccess(number):
        global connected,cap
        if number==0:
            connected=False
            tk.messagebox.showerror(lang[112],lang[113]) #Connection Failed
        else:
            connected=True
            connectbutton.configure(text=lang[87],fg_color='dark red',hover_color='#530000',command=lambda : asyncio.run(stream_reset())) #Connection Success
            calibratebutton.configure(state="normal")
            previewbutton.configure(state="normal")
            recordbutton.configure(state="normal")
            
    async def connect():
        global tobiiglasses,disconnect
        try:
            tobiiglasses = await g3pylib.connect_to_glasses.with_zeroconf(True,2000)
            connectsuccess(1)
            await batteryrefresh()
            disconnect = False
            previewavailfun()
            await tobiiglasses.close()
        except:
            connectsuccess(0)

    #Live Preview Functions
    def video_stop():
        global stopper
        stopper = True
        previewbutton.configure(text=lang[88],command=video,fg_color=("#3b8ed0","#1f6aa5"),hover_color=("#36719f","#144870")) #Start Live View

    async def popvideo_thread():
        global disconnect
        async with g3pylib.connect_to_glasses.with_zeroconf(True,2000) as g3:
            async with g3.stream_rtsp() as streams:
                async with streams.scene_camera.decode() as decoded_stream:
                    previewbutton.configure(text=lang[91],command=video_stop,fg_color="dark red",hover_color='#530000',state="normal") #Stop Live View
                    previewpopfun()
                    while True:
                        frame, _timestamp = await decoded_stream.get()
                        image = frame.to_ndarray(format="bgr24")
                        cv2.imshow("Tobii Glasses Live View", image)
                        if (cv2.waitKey(1)!=-1 or disconnect or stopper):
                            break
                    cv2.destroyAllWindows()
                    decoded_stream.task_done()
                    await g3.close()
                    video_stop()
                    if disconnect:
                        connectpreviewfun()
                    else:
                        previewavailfun()
                    
    def popvideo_wrapper():
        asyncio.run(popvideo_thread())
        
    def video():
        global stopper
        stopper = False
        previewloadfun()
        previewbutton.configure(text=lang[92],state='disabled') #Loading
        if popenabled:
            thread1 = Thread(target=popvideo_wrapper)
            thread1.start()
        else:
            thread2 = Thread(target=invideo_wrapper)
            thread2.start()

    def invideo_wrapper():
        asyncio.run(invideo_thread())

    async def invideo_thread():
        global disconnect #Disconnect Flag
        async with g3pylib.connect_to_glasses.with_zeroconf(True,2000) as g3:
            async with tobiiglasses.stream_rtsp() as streams:
                async with streams.scene_camera.decode() as cap:
                    previewbutton.configure(text=lang[91],command=video_stop,fg_color="dark red",hover_color='#530000',state="normal")
                    clearlive()
                    videolabel.grid(row=1,column=0,rowspan=4,padx=20,pady=(10,20),sticky='nw')
                    while True:
                        frame, timestamp_ = await cap.get()
                        frame = frame.to_ndarray(format="rgb24")
                        resized = cv2.resize(frame, (newx,newy))
                        img = Image.fromarray(resized)
                        imgtk = ImageTk.PhotoImage(image=img)
                        videolabel.configure(image=imgtk)
                        videolabel.imgtk = imgtk
                        if (disconnect or stopper):
                            break
                    cap.task_done()
                    await g3.close()
                    video_stop()
                    videolabel.grid_forget()
                    if disconnect:
                        connectpreviewfun()
                    else:
                        previewavailfun()

    #Record
    async def record():
        #tk.messagebox.showinfo("Code under development","The code for this function is not yet written. Please wait for a few days and check your version to make sure you are on the latest build.")
        global recordactive
        recordbutton.configure(text=lang[92],state="disabled") #Loading
        g2 = await g3pylib.connect_to_glasses.with_zeroconf(True,2000)
        res = await g2.recorder.start()
        await g2.close()
        if res:   
            recordactive=True
            text20.grid_forget()
            recordstatus.configure(image=imgtk26)
            recordstatus.grid(row=0,column=0,padx=(20,0),pady=(10,0),sticky='w')
            recordbutton.configure(text=lang[93],command=lambda: asyncio.run(stoprecord()),state="normal") #Stop Rceording
        else:
            tk.messagebox.showerror(lang[94],lang[95])
            recordactive=False
            recordbutton.configure(text=lang[84],state="normal",command=lambda: asyncio.run(record()))

    def stophelper():
        recordstatus.grid_forget()
        text20.grid(row=0,column=0,padx=(20,0),pady=(10,0),sticky='w')

    async def stoprecord():
        global recordactive
        recordactive=False
        recordbutton.configure(text=lang[92],state="disabled") #Loading
        g4 = await g3pylib.connect_to_glasses.with_zeroconf(True,2000)
        res = await g4.recorder.stop()
        await g4.close()
        if res:
            recordbutton.configure(text=lang[84],state="normal",command=lambda: asyncio.run(record()))
            recordstatus.configure(image=imgtk27)
            recordstatus.after(3000,stophelper)

        else:
            tk.messagebox.showerror(lang[96],lang[97]) #Error Stopping Recording
            stophelper()
            recordbutton.configure(text=lang[93],state="normal",command=lambda: asyncio.run(stoprecord()))

    #Calibrate
    async def calibrate():
        tk.messagebox.showinfo(lang[98],lang[99]) #Calibrate Message
        tobiiglasses = await g3pylib.connect_to_glasses.with_zeroconf(True,2000)
        res = await tobiiglasses.calibrate.run()
        await tobiiglasses.close()
        if res==False:
            calibratebutton.configure(text=lang[100],fg_color="dark red",state="disabled") #Calibration Failed
            tk.messagebox.showwarning(lang[101],lang[102]) #Calibration Failed Message
            calibratebutton.configure(text=lang[81],fg_color=("#3b8ed0","#1f6aa5"),state="normal",hover_color=("#36719f","#144870")) #Calibrate Glasses
        else:
            calibratebutton.configure(text=lang[103],fg_color="green",state="normal",hover_color="#006400") #Calibrate Again

    #Disconnect and Live Preview Helper Functions
    async def stream_reset():
        global recordactive,disconnect,connected
        disconnect = True
        if recordactive==False:
            connected=False
            stophelper()
            text20.grid(row=0,column=0,padx=(20,0),pady=(10,0),sticky='w')
            connectbutton.configure(text=lang[78],fg_color=("#3b8ed0","#1f6aa5"),hover_color=("#36719f","#144870"),state="normal",command=lambda : asyncio.run(connect())) #Connect Glasses
            calibratebutton.configure(state="disabled",text=lang[81],fg_color=("#3b8ed0","#1f6aa5"),hover_color=("#36719f","#144870")) #Calibrate Glasses
            recordbutton.configure(state="disabled",text=lang[84],command=record) #Record Now
            previewbutton.configure(text=lang[88],command=video,fg_color=("#3b8ed0","#1f6aa5"),hover_color=("#36719f","#144870"),state="disabled") #Start Live View
            text21.grid_forget()
            text22.grid_forget()
            batteryicon.grid_forget()
            timelefticon.grid_forget()
            videolabel.grid_forget()
            connectpreviewfun()      
        else:
            tk.messagebox.showerror(lang[114],lang[115]) #Recording Still in Progress

    def resizerecord(e):
        def resize_images():
            global imgtk56, imgtk57, imgtk58, newx, newy
            newx = int(e.width-40)
            newy= int((9 / 16) * (e.width-40))
            imgtk56 = CTkImage(light_image=image27, size=(newx, newy))
            imgtk57 = CTkImage(light_image=image28, size=(newx, newy))
            imgtk58 = CTkImage(light_image=image29, size=(newx, newy))
            videoframe.after(100, update_ui)
        Thread(target=resize_images).start()

    def update_ui():
        connectpreview.configure(image=imgtk56)
        previewpop.configure(image=imgtk57)
        previewavail.configure(image=imgtk58)
        previewload.configure(width=newx,height=newy)

    def clearlive():
        previewload.unload()
        previewavail.grid_forget()
        connectpreview.grid_forget()
        previewload.grid_forget()
        previewpop.grid_forget()

    def previewpopfun():
        previewload.unload()
        previewavail.grid_forget()
        connectpreview.grid_forget()
        previewload.grid_forget()
        previewpop.grid(row=1,column=0,columnspan=5,padx=20,pady=(5,10),sticky='nsew')

    def connectpreviewfun():
        previewavail.grid_forget()
        previewload.grid_forget()
        connectpreview.grid(row=1,column=0,columnspan=5,padx=20,pady=(5,10),sticky='nsew')
        previewpop.grid_forget()

    def previewavailfun():
        previewload.grid_forget()
        previewavail.grid(row=1,column=0,columnspan=5,padx=20,pady=(5,10),sticky='nsew')
        connectpreview.grid_forget()
        previewpop.grid_forget()

    def previewloadfun():
        thread3 = Thread(target=previewloadfun_thread)
        thread3.start()

    def previewloadfun_thread():
        previewload.load(".\Resources\Loading1.gif")
        previewavail.grid_forget()
        connectpreview.grid_forget()
        previewpop.grid_forget()
        previewload.grid(row=1,column=0,columnspan=5,padx=20,pady=(5,10),sticky='nsew')

    #Analyze Page Functions
    def analyze_reset():
        global exportstatus,viewadvancer,viewadvancer1,table1,table2,table3,framebuttons
        exportstatus = 3
        export_reset()
        progressframe.grid(row=7,column=0,columnspan=3,sticky='nsew',padx=20,pady=20)
        text85.configure(text=lang[63],text_color=("black","white"))
        text86.configure(text="0%")
        progressbar.set(0)
        export1.configure(state='disabled')
        export2.configure(state='disabled')
        export5.configure(state='disabled')
        imgframe.grid_forget()
        imgframe2.grid_forget()
        imgframe3.grid_forget()
        showgraph.grid_forget()
        deadmanbutton.grid_forget()
        violateradar.grid_forget()
        editnames.grid_forget()
        detectionsbutton.grid_forget()
        clearbutton.grid_forget()
        imgframe4.grid_forget()
        imgframe5.grid_forget()
        imgframe6.grid_forget()
        
        for i in imgframe.grid_slaves(row=None,column=None):
            i.grid_remove()
        try:
            for i in imgframe2.grid_slaves(row=None,column=None):
                i.grid_remove()
            for i in imgframe5.grid_slaves(row=None,column=None):
                i.grid_remove()
        except:
            pass
        try:
            table1.grid_forget()
            table2.grid_forget()
            table3.grid_forget()
            sct.grid_forget()
        except:
            pass
        
        viewadvancer = 0
        viewadvancer1 = 0
        viewdata.configure(state='disabled')
        if compact==1:
            viewdata.configure(fg_color="#333333")
            framebuttons = [settingsbutton,recorddata,exportdata,analyzedata]
        playpausebutton.configure(state='disabled')
        forwardbutton.configure(state='disabled')
        backbutton.configure(state='disabled')
        viewslider.configure(state='disabled')
        viewend.configure(text="0:00")
        player.grid_forget()
        viewtext.pack_forget()
        text44.configure(text="\n"*46)
        text44.grid(row=0,column=0,columnspan=3,padx=(10,0),pady=(10,0),sticky='w')

    def loadjson():
        global loaddata,loadeddata,loadfolder
        analyze_reset()
        #Main code disable only for faster access
        loadfile.configure(fg_color=("#3b8ed0","#1f6aa5"),text='Load File',hover_color=("#36719f","#144870"))
        text13.configure(text=lang[62]) #Load SD Folder
        loaddata = tk.filedialog.askdirectory()
        if loaddata!='':
            if not os.path.exists(loaddata+'\scenevideo.mp4') and not os.path.exists(loaddata+'\gazedata.gz') and not os.path.exists(loaddata+'\imudata.gz'):
                tk.messagebox.showerror(lang[116],lang[117]) #Problem Loading Files
                return 0
            loadfolder = loaddata
            loaddata = extract(loaddata)
            loadeddata = True
            analyze.configure(state="normal")
            text43.configure(text=lang[64],text_color=("black","white"))
            loadfile.configure(fg_color='green',text='Loaded',hover_color="#006400")
            t=lang[105]+loadfolder #Using 
            if len(t)>90:
                t=lang[104] #Project Folder Loaded
            text13.configure(text=t)
            value[15]=loaddata+"\n"
            writesettings()
        else:
            loadeddata = False
            analyze.configure(state="disabled")
            text43.configure(text=lang[106],text_color=("red","yellow")) #Load Project Folder

    def autoloadfiles():
        global loaddata,loadeddata,value,loadfolder
        temp1=value[15][:-1]
        if rememberloadbool==False:
            text43.configure(text=lang[106],text_color=("red","yellow")) #Load Project Folder
        if os.path.exists(temp1):
            pass
        else:
            temp1='Load1'
            text43.configure(text=lang[106],text_color=("red","yellow"))
        if rememberloadbool==True and temp1!='Load1':
            loaddata = temp1
            loadeddata = True
            analyze.configure(state="normal")
            loadfile.configure(fg_color='green',text='Auto Loaded',hover_color="#006400")
            loadfolder = os.path.dirname(loaddata)
            t1=lang[105]+loadfolder
            if len(t1)>90:
                t1=lang[104]
            text13.configure(text=t1)
            text43.configure(text=lang[64],text_color=("black","white"))

    def uniqueduration():
        global notimeslooked,totduration
        notimeslooked=[0 for i in range(len(UAOIName))]
        totduration=[0 for i in range(len(UAOIName))]
        for i in range(len(UAOIName)):
            for j in range(len(AOIName)):
                if UAOIName[i]==AOIName[j]:
                    notimeslooked[i]=notimeslooked[i]+1
                    totduration[i]=totduration[i]+(EndSec[j]-StartSec[j])

    def editAOI():
        if os.path.exists(loadfolder+r"\names.txt"):
            f = open(loadfolder+r"\names.txt","r")
            store = f.readlines()
            f.close()
            if len(store)==unique_count: #Check for Corruption
                ans = tk.messagebox.askyesno(lang[160],lang[161]) #Confirm Previous Load
                if ans:
                    for i in range(len(store)):
                        store[i] = store[i][:-1] #Remove \n
                    saveAOI('',store)
                    return 0
        root.bind("<Return>",saveAOI)
        editnames.configure(text=lang[107],command=saveAOI)
        col=0
        row=2
        for i in range(len(UAOIName)):
            globals()['t%s' % i].grid_forget()
            globals()['t%s' % i] = CTkEntry(imgframe2,placeholder_text=UAOIName[i],
                                width=100,
                                height=30,
                                border_width=2,
                                corner_radius=5,
                                font=CTkFont(size=15),
                                justify=CENTER)
            if col==4:
                col=0
                row=row+2
            globals()['t%s' % i].grid(row=row,column=col,sticky='n',pady=(0,10)) 
            col=col+1

    def saveAOI(a='',override=[]):
        global UAOIName,AOIName,pos,picarray,table1,table2,table3
        root.unbind("<Return>")
        store=[]
        check=[]
        if override==[]:
            for i in range(len(UAOIName)):
                a=(globals()['t%s' % i].get()) #Get All Inputs
                if a=='':
                    store.append(UAOIName[i])
                else:
                    store.append(a)
        else:
            store = override
        if use_manual==0: #Check if duplicates entered
            flag=0
            for i in store:
                if i in check:
                    tk.messagebox.showwarning(lang[118],lang[119]) #Failed to Assign Unique Names
                    flag=1
                    break
                else:
                    check.append(i)
            if flag==1: #Cancel Operation and Revert to Edit
                return 0
            for i in range(len(AOIName)): #Proceed to Update Lists
                for j in range(len(UAOIName)):
                    if AOIName[i]==UAOIName[j]:
                        AOIName[i]=store[j]
            UAOIName=store
        else:
            AOIName=store
            pos.clear()
            for i in range(len(store)):
                if store[i] not in check:
                    check.append(store[i])
                    pos.append(i)
            UAOIName=check
            editnames.configure(state='disabled')

        #Save this Configuration
        save = store
        f = open(loadfolder+r"\names.txt","w")
        for i in save:
            f.write(i)
            f.write("\n")
        f.close()
        
        if use_manual==0:
            col=0
            row=2
            for i in range(len(UAOIName)):
                if col==4:
                    col=0
                    row=row+2
                globals()['t%s' % i].grid_forget()
                globals()['t%s' % i] = CTkLabel(imgframe2,text=UAOIName[i],font=CTkFont(size=15))
                globals()['t%s' % i].grid(row=row,column=col,sticky='n',pady=(0,10)) 
                col=col+1
        else:
            for i in imgframe2.grid_slaves(row=None,column=None):
                i.grid_remove()
            text45.grid(row=0,column=0,padx=(10,0),pady=10,sticky='w')
            ucounttext=lang[108]+str(len(UAOIName))
            text45.configure(text=ucounttext)

            col=0
            row=1
            for i in range(len(picarray)):
                if i in pos:
                    img = Image.fromarray(picarray[i])
                    imgsave.append(img)
                    aimgtk = CTkImage(light_image=img,size=(int(1920*0.17),int(1080*0.17)))
                    if col==4:
                        col=0
                        row=row+2
                    e1 = CTkLabel(imgframe2,text='',image=aimgtk)
                    e1.grid(row=row,column=col,sticky='nsew',padx=(10,0),pady=(0,5))
                    col=col+1
            
            col=0
            row=1
            for i in range(len(UAOIName)):
                if col==4:
                    col=0
                    row=row+2
                globals()['t%s' % i] = CTkLabel(imgframe2,text=UAOIName[i],font=CTkFont(size=15))
                globals()['t%s' % i].grid(row=row+1,column=col,sticky='nsew',pady=(0,10))
                col=col+1

            uniqueduration()

        col=0
        row=2
        for i in range(len(AOIName)):
            if col==4:
                col=0
                row=row+2
            e2 = CTkLabel(imgframe,text=AOIName[i],font=CTkFont(size=15))
            e2.grid(row=row,column=col,sticky='nsew',pady=(0,10))    
            col=col+1
        editnames.configure(text=lang[71],command=editAOI) #Edit Names
        viewtext.configure(text=f" {AOIName[viewadvancer]} | {headtable[viewadvancer1][0]} ")
        table1.grid_forget()
        table2.grid_forget()
        table3.grid_forget()
        showtables()

    def analyzefun():
        global AOIName,UAOIName,StartSec,EndSec,GazeX,GazeY,Timestamps,color,size,unique_count,imgsave,pos,picarray,exportstatus,headtable,GazeXmin,GazeYmin,viewadvancer,viewadvancer1,headmode,useai,atolbox,framebuttons
        imgsave=[]
        analyze_reset()
        if os.path.exists(loaddata)==False:
            tk.messagebox.showerror(lang[120],lang[121]) #Opening Files Failed
            return 0
        try:
            count,unique_count,picarray,answerarray,answer,pos,AOIName,UAOIName,StartSec,EndSec,GazeX,GazeY,Timestamps,color,size,headtable,headduration,violation,imageai_avail=analyzer(loaddata,loadfolder+"\scenevideo.mp4",asense,useai,use_manual,thit,headmode,headsen,atolbox) 
        except PermissionError:
            tk.messagebox.showerror(lang[122],lang[123]) #Permission Error
            text85.configure(text=lang[122],text_color=("red","yellow"))
            return 0
        except:
            tk.messagebox.showerror(lang[124],lang[125]) #Failed to Analyze
            text85.configure(text=lang[124],text_color=("red","yellow"))
            return 0
        
        #Lite Version Check
        if useai!=0 and imageai_avail==False:
            tk.messagebox.showwarning(lang[154],lang[155])
            useai=0
            updateuseai(True)
        
        #Memory Overload Check
        if len(picarray)>400:
            tk.messagebox.showwarning(lang[147],lang[148]) #Too much areas of interest

        imgframe.configure(height=270)
        imgframe2.configure(height=270)
        imgframe5.configure(height=40)
        counttext=lang[109]+str(count)
        ucounttext=lang[108]+str(unique_count)
        text44.grid(row=0,column=0,padx=(10,0),pady=10,sticky='w')
        text44.configure(text=counttext)
        if useai in [1,2]:
            UAOIName.clear()
            store=[]
            for i in range(len(answer)):
                if i in pos:
                    if answer[i] not in store:
                        store.append(answer[i])
                        UAOIName.append(answer[i]+' 1')
                    else:
                        l=store.count(answer[i])
                        UAOIName.append(answer[i]+f' {l+1}')
            store.clear()
            for i in range(len(AOIName)):
                word = AOIName[i].split()
                store.append(UAOIName[int(word[1])-1])
            AOIName=store

        if useai==3:
            if use_manual==1:
                UAOIName=answer
            else:
                UAOIName.clear()
                store=[]
                for i in answer:
                    if i not in store:
                        store.append(i)
                UAOIName = store
            AOIName = answer
        
        progressframe.grid_forget()
        imgframe.grid(row=7,column=0,columnspan=3,sticky='nsew',padx=20,pady=20)
        col=0
        row=1
        for i in range(len(picarray)):
            if useai in [1,2,3]:
                img = Image.fromarray(answerarray[i])
            else:
                img = Image.fromarray(picarray[i])
            aimgtk = CTkImage(light_image=img,size=(int(1920*0.17),int(1080*0.17)))
            if col==4:
                col=0
                row=row+2
                h = imgframe.cget("height")
                imgframe.configure(height=h+240)
            e1 = CTkLabel(imgframe,text='',image=aimgtk)
            e1.grid(row=row,column=col,sticky='nsew',padx=(10,0),pady=(0,5))
            e2 = CTkLabel(imgframe,text=AOIName[i],font=CTkFont(size=15))
            e2.grid(row=row+1,column=col,sticky='nsew',pady=(0,10))
            col=col+1
        
        imgframe2.grid(row=8,column=0,columnspan=3,sticky='nsew',padx=20)
        text45.grid(row=0,column=0,padx=(10,0),pady=10,sticky='w')
        text45.configure(text=ucounttext)

        #Dead Man Check
        flag1,flag2 = False,False
        for i in range(len(AOIName)):
            if (round(EndSec[i]-StartSec[i],2)>adead):
                flag1 = True
                break
        for i in headduration:
            if (i>adead):
                flag2 - True
                break
        if flag1 and flag2:
            deadmanbutton.configure(fg_color='dark red',hover_color='#530000',command=lambda :tk.messagebox.showwarning(lang[51],lang[110]))
        else:
            deadmanbutton.configure(fg_color=("#3b8ed0","#1f6aa5"),hover_color=("#36719f","#144870"),command=lambda :tk.messagebox.showinfo(lang[51],lang[73]))

        col=0
        row=1
        for i in range(len(picarray)):
            if i in pos:
                img = Image.fromarray(picarray[i])
                imgsave.append(img)
                aimgtk = CTkImage(light_image=img,size=(int(1920*0.17),int(1080*0.17)))
                if col==4:
                    col=0
                    row=row+2
                    h = imgframe2.cget("height")
                    imgframe2.configure(height=h+240)
                e1 = CTkLabel(imgframe2,text='',image=aimgtk)
                e1.grid(row=row,column=col,sticky='nsew',padx=(10,0),pady=(0,5))
                col=col+1

        col=0
        row=1
        for i in range(unique_count):
            if col==4:
                col=0
                row=row+2
            globals()['t%s' % i] = CTkLabel(imgframe2,text=UAOIName[i],font=CTkFont(size=15))
            globals()['t%s' % i].grid(row=row+1,column=col,sticky='nsew',pady=(0,10))
            col=col+1
        
        
        text50.grid(row=0,column=0,columnspan=3,sticky="w",pady=(10,0),padx=(10,0))
        headup.configure(text=f"{lang[68]}\n{headduration[0]}s\n\n({round((headduration[0]/(headduration[0]+headduration[1]+headduration[2]))*100,2)}%)")
        headup.grid(row=1,column=0,pady=(0,20))
        headlevel.configure(text=f"{lang[69]}\n{headduration[1]}s\n\n({round((headduration[1]/(headduration[0]+headduration[1]+headduration[2]))*100,2)}%)")
        headlevel.grid(row=1,column=1,pady=(0,20))
        headdown.configure(text=f"{lang[70]}\n{headduration[2]}s\n\n({round((headduration[2]/(headduration[0]+headduration[1]+headduration[2]))*100,2)}%)")
        headdown.grid(row=1,column=2,pady=(0,20))
        export1.configure(state='normal')
        export2.configure(state='normal')
        export5.configure(state='normal')
        exportstatus = 0

        imgframe6.grid(row=9,column=0,columnspan=3,sticky='nsew',padx=20,pady=20)
        editnames.grid(row=0,column=0,sticky='w',padx=(0,10))
        detectionsbutton.grid(row=0,column=1,sticky='w',padx=(0,10))
        clearbutton.grid(row=0,column=2,sticky='w')
        imgframe4.grid(row=10,column=0,columnspan=3,sticky='nsew',padx=20,pady=(0,20))
        imgframe3.grid(row=12,column=0,columnspan=3,sticky='nsew',padx=20,pady=20)
        showgraph.grid(row=0,column=0,sticky='w',padx=(0,10))
        violateradar.grid(row=0,column=1,sticky='w',padx=(0,10))
        deadmanbutton.grid(row=0,column=2,sticky='w')

        imgframe5.grid(row=13,column=0,columnspan=3,sticky='nsew',padx=20,pady=(0,20))
        text51.configure(text=f"{lang[72]}: {len(violation)}")
        text51.grid(row=0,column=0,columnspan=3,sticky="w",pady=(10,10),padx=(10,0))
        col=0
        row=1
        if len(violation)!=0:
            imgframe5.configure(height=270)
            for i in range(len(violation)):
                if i in pos:
                    img = Image.open(violation[i])
                    imgsave.append(img)
                    aimgtk = CTkImage(light_image=img,size=(int(1920*0.17),int(1080*0.17)))
                    if col==4:
                        col=0
                        row=row+2
                        h = imgframe5.cget("height")
                        imgframe5.configure(height=h+240)
                    e1 = CTkLabel(imgframe5,text='',image=aimgtk)
                    e1.grid(row=row,column=col,sticky='nsew',padx=(10,0),pady=(0,10))
                    col=col+1
        
        if unique_count==0:
            editnames.configure(state='disabled')
            showgraph.configure(state='disabled')
        else:
            editnames.configure(state='normal')
            showgraph.configure(state='normal')
        uniqueduration()
        showtables()
        if len(AOIName)!=0:
            playpausebutton.configure(state='normal')
            forwardbutton.configure(state='normal')
            backbutton.configure(state='normal')
            viewslider.configure(state='normal')
            viewadvancer = 0
            viewadvancer1 = 0
            temp = []
            GazeXmin,GazeYmin = [],[]
            for i in range(len(Timestamps)):
                if i%100==0:
                    GazeXmin.append(GazeX[i])
                    GazeYmin.append(GazeY[i])
                    temp.append(Timestamps[i])
            Timestamps = temp
            if exp:
                img = Image.fromarray(picarray[0])
                tk_image = CTkImage(light_image=img, size=(192,108))
                viewtext.configure(text=f" {AOIName[0]} ", image=tk_image, compound="top")  
            else:
                viewtext.configure(text=f" {AOIName[0]} ")
            player.grid(row=0,column=0,columnspan=5,sticky='nsew',padx=10,pady=10)
            player.load(loadfolder+"\scenevideo.mp4")
            viewdata.configure(state='normal')
            if compact==1:
                viewdata.configure(fg_color=("#3b8ed0","#1f6aa5"))
                framebuttons = [settingsbutton,recorddata,exportdata,analyzedata,viewdata]
            violateradar.configure(state="normal")
            viewtext.configure(text=f" {AOIName[viewadvancer]} | {headtable[viewadvancer1][0]} ")
        else:
            violateradar.configure(state="disabled")

    def update_progress():
        try:
            progress_value = progress_queue.get_nowait()
            progressbar.set(progress_value/100)
            text86.configure(text=str(progress_value)+"%")
            if (progress_value==5):
                text85.configure(text=lang[49])
            elif (progress_value==10):
                text85.configure(text=lang[48])
            elif (progress_value==15):
                text85.configure(text=lang[77])
            elif (progress_value==25):
                text85.configure(text=lang[78])
            elif (progress_value==40):
                text85.configure(text=lang[79])
            elif (progress_value==50):
                text85.configure(text=lang[80])
            elif (progress_value==60):
                text85.configure(text=lang[82])
            elif (progress_value==70):
                text85.configure(text=lang[83])
            elif (progress_value==75):
                text85.configure(text=lang[162])
            elif (progress_value==90):
                text85.configure(text=lang[80])
            elif (progress_value==95):
                text85.configure(text=lang[163])

            if progress_value < 100:
                # Schedule the update function again after 100 milliseconds
                root.after(100, update_progress)
        except Empty:
            # If the queue is empty, check again after 100 milliseconds
            root.after(100, update_progress)

    def analyze_wrapper():
        time.sleep(0.5)
        analyze.configure(text=lang[54],state="disabled")
        loadfile.configure(state="disabled")
        try:
            analyzefun()
            global progress
            progress = 100
            progress_queue.put(progress)
            progressbar.set(1)
            analyze.configure(text=lang[66],state="normal")
            loadfile.configure(state="normal")
            return
        except _thread.exit:
            return

    def start_analysis():
        global thread1
        thread1 = _thread.start_new_thread(analyze_wrapper, ())
        #thread1.start()
        root.after(100,update_progress)

    def showtables():
        global table1,table2,table3,sct,styletree
        styletree = ttk.Style(analyzeframe)
        styletree.theme_use("clam")
        if value[2][:-1]=="Dark":
            styletree.configure("Treeview", background="#333333",fieldbackground="#333333", foreground="white")
        else:
            styletree.configure("Treeview", background="white",fieldbackground="white", foreground="black")
        table2 = ttk.Treeview(analyzeframe)
        table2.grid_forget()
        table2['columns'] = ("AOI","Start","End","Duration","Type")
        table2.column("#0",width=50)
        table2.column("AOI",anchor=W,width=135)
        table2.column("Start",anchor=CENTER,width=90)
        table2.column("End",anchor=CENTER,width=90)
        table2.column("Duration",anchor=CENTER,width=90)
        table2.column("Type",anchor=W,width=100)
        table2.heading("#0",text="No",anchor=CENTER)
        table2.heading("AOI",text="Area of Interest",anchor=W)
        table2.heading("Start",text="Start Time (s)",anchor=CENTER)
        table2.heading("End",text="End Time (s)",anchor=CENTER)
        table2.heading("Duration",text="Duration (s)",anchor=CENTER)
        table2.heading("Type",text="Description",anchor=W)

        id=1
        for i in range(len(AOIName)):
            if round(EndSec[i]-StartSec[i],2)<(thit/100)+1:
                desc='Short'
            elif round(EndSec[i]-StartSec[i],2)<(thit/100)+2:
                desc='Medium'
            elif round(EndSec[i]-StartSec[i],2)<(thit/100)+5:
                desc='Long'
            else:
                desc='Prolonged'
            table2.insert(parent='',index='end',iid=id,text=id,values=(AOIName[i],f"{round(StartSec[i],2)}",f"{round(EndSec[i],2)}",f"{round(EndSec[i]-StartSec[i],2)}",desc))
            id=id+1
        table2.grid(row=11,column=0,padx=(20,0),pady=0,sticky='nsew')

        table1 = ttk.Treeview(analyzeframe)
        table1.grid_forget()
        table1['columns'] = ("AOI","No","Duration")
        table1.column("#0",width=50)
        table1.column("AOI",anchor=W,width=140)
        table1.column("No",anchor=CENTER,width=120)
        table1.column("Duration",anchor=CENTER,width=90)
        table1.heading("#0",text="No",anchor=CENTER)
        table1.heading("AOI",text="Unique Area of Interest",anchor=W)
        table1.heading("No",text="No of times looked",anchor=CENTER)
        table1.heading("Duration",text="Duration (s)",anchor=CENTER)

        id=1
        for i in range(len(UAOIName)):
            table1.insert(parent='',index='end',iid=id,text=id,values=(UAOIName[i],notimeslooked[i],f"{round(totduration[i],2)}"))
            id=id+1
        table1.grid(row=11,column=1,padx=(10,10),pady=0,sticky='nsew')

        table3 = ttk.Treeview(analyzeframe)
        table3.grid_forget()
        table3['columns'] = ("Type","Start","End","Duration")
        table3.column("#0",width=50)
        table3.column("Type",anchor=W,width=140)
        table3.column("Start",anchor=CENTER,width=90)
        table3.column("End",anchor=CENTER,width=90)
        table3.column("Duration",anchor=CENTER,width=90)
        table3.heading("#0",text="No",anchor=CENTER)
        table3.heading("Type",text="Head Orientation",anchor=W)
        table3.heading("Start",text="Start Time (s)",anchor=CENTER)
        table3.heading("End",text="End Time (s)",anchor=CENTER)
        table3.heading("Duration",text="Duration (s)",anchor=CENTER)

        id=1
        for i in range(len(headtable)):
            table3.insert(parent='',index='end',iid=id,text=id,values=(headtable[i][0],round(headtable[i][1],2),round(headtable[i][2],2),round(headtable[i][2]-headtable[i][1],2)))
            id=id+1
        table3.grid(row=11,column=2,padx=(0,20),pady=0,sticky='nsew')

    def showgraphs():
        plt.ioff
        f = plt.figure()
        f.set_figwidth(15)
        f.set_figheight(8)
        plt.subplot(2,2,1)
        plt.scatter(GazeX,GazeY,s=size,c=color,alpha=0.5,cmap='Reds')
        plt.xlabel('Gaze X 2D Coordinates')
        plt.ylabel('Gaze Y 2D Coordinates')
        cbar = plt.colorbar()
        cbar.set_label('Fixation')

        plt.subplot(2,2,2)
        plt.bar(UAOIName,notimeslooked,width=0.3)
        plt.xlabel('Areas of Interest')
        plt.ylabel('No of times looked')

        plt.subplot(2,2,3)
        plt.plot(StartSec,AOIName,'o',color='Red')
        plt.ylabel('Areas of Interest')
        plt.xlabel('Time(s)')

        plt.subplot(2,2,4)
        plt.bar(UAOIName,totduration,width=0.3,color='Orange')
        plt.xlabel('Areas of Interest')
        plt.ylabel('Total Duration(s)')
        plt.show()

    def radar():
        global aviolate
        flag = False
        check = [0,]
        violation = []
        violation_type = []
        for i in range(len(AOIName)):
            if AOIName[i]=="Radar" and (StartSec[i]-check[-1])<aviolate:
                check.append(StartSec[i])
            elif AOIName[i]=="Radar" and (StartSec[i]-check[-1])>aviolate:
                violation.append(StartSec[i])
                violation_type.append("Looked at Radar too late")
                check.append(StartSec[i])
            elif AOIName[i]!="Radar" and (StartSec[i]-check[-1])>aviolate:
                violation_type.append("Failed to look at Radar")
                violation.append(StartSec[i])
                check.append(StartSec[i])
            else:
                pass
        
        if len(violation)==0:
            tk.messagebox.showinfo(lang[126],lang[127]) #No Radar Violations
        else:
            violations = "\n"
            for i in range(len(violation)):
                violations=violations+"At "+str(round(violation[i],2))+"s : "+violation_type[i]+"\n"
            tk.messagebox.showwarning(lang[128],f"{lang[129]}{violations}")
    
    def nogozonefun():
        dir = os.getcwd() + r"\NoZones"
        os.system(f'start "" "{dir}"')

    def detectionsfun():
        dir = loadfolder+r"\detections"
        os.system(f'start "" "{dir}"')

    def clearbuttonfun():
        analyzeframe.event_generate("<MouseWheel>", delta=1000*120)
        root.after(10,analyze_reset)


    #Export Page Functions
    def exportsaver():
        global exportsavefolder
        export_reset()
        ask=tk.filedialog.askdirectory()
        if ask=="" and exportsavefolder=='Expsave':
            exportsavefolder=os.getcwd()
        elif ask=="" and exportsavefolder!="Expsave":
            pass
        else:
            exportsavefolder=ask
        text54.configure(text=f"{lang[105]}{exportsavefolder}") #Using
        exportsave.configure(text=lang[111],fg_color='green',hover_color="#006400") #Selected
        value[19]=exportsavefolder+'\n'
        writesettings()

    def exportcsv():
        global exportsavefolder,exportstatus
        export_reset()
        if exportsavefolder=='Expsave':
            exportsavefolder=os.getcwd()
            text54.configure(text=f"{lang[105]}{exportsavefolder}")
            value[19]=exportsavefolder+'\n'
            writesettings()
            if askexport==0:
                tk.messagebox.showinfo(lang[130],f"{lang[131]}{exportsavefolder}")
        
        c=1
        path=exportsavefolder+"/"+f'EyePort Scenario {c}.csv'
        while os.path.exists(path):
            c=c+1
            path=exportsavefolder+"/"+f'EyePort Scenario {c}.csv'
        if c!=1 and overexport==1:
            path=exportsavefolder+"/"+f'EyePort Scenario {c-1}.csv'
            d=c-1
        else:
            d=c

        if askexport==1:
            askpath=tk.filedialog.asksaveasfilename(initialdir=exportsavefolder,initialfile=f'EyePort Scenario {d}.csv',filetypes=[("CSV File","*.csv")])
            if askpath!='':
                if askpath[-4:]!='.csv':
                    askpath=askpath+'.csv'
                path=askpath
        
        try:
            fcsv=open(path,'w',newline='')
            writer = csv.writer(fcsv)
            writer.writerow(['Time','ActiveFunction','ActiveFunctionOutput','DownstreamCoupledFunction','CoupledFunctionAspect','TimeTolerance'])
            
            for i in range(len(AOIName)):
                if i==len(AOIName)-1:
                    writer.writerow([int(StartSec[i]),UAOIName.index(AOIName[i])+1,'Observing '+AOIName[i],0,'I'])
                    writer.writerow([int(EndSec[i]),0,'Changing Focus or Saccading',0,'I'])
                    break
                else:
                    writer.writerow([int(StartSec[i]),UAOIName.index(AOIName[i])+1,'Observing '+AOIName[i],0,'I'])
                    writer.writerow([int(EndSec[i]),0,'Changing Focus or Saccading',UAOIName.index(AOIName[i+1])+1,'I'])
            fcsv.close()
            exportstatus = 1
            exportsuccessful()
            print("Exported CSV File")
        except PermissionError:
            exportstatus = 2
            exportfailed()
            tk.messagebox.showerror(lang[122],lang[132])
            return 0
        except:
            exportstatus = 2
            exportfailed()
            tk.messagebox.showerror(lang[133],lang[134])
            return 0

    def exportxmfv():
        global exportsavefolder,exportstatus
        export_reset()
        if exportsavefolder=='Expsave':
            exportsavefolder=os.getcwd()
            text54.configure(text=f"{lang[105]}{exportsavefolder}")
            value[19]=exportsavefolder+'\n'
            writesettings()
            if askexport==0:
                tk.messagebox.showinfo(lang[130],f"{lang[131]}{exportsavefolder}")
        
        try:
            creator(UAOIName, exportsavefolder, overexport, askexport)
            exportstatus = 1
            exportsuccessful()
            print("Exported XMFV File")
        except PermissionError:
            exportstatus = 2
            exportfailed()
            tk.messagebox.showerror(lang[122],lang[132])
            return 0
        except:
            exportstatus = 2
            exportfailed()
            tk.messagebox.showerror(lang[133],lang[134])
            return 0

    def exportxlsx():
        global exportsavefolder,exportstatus
        export_reset()
        if exportsavefolder=='Expsave':
            exportsavefolder=os.getcwd()
            text54.configure(text=f"{lang[105]}{exportsavefolder}")
            value[19]=exportsavefolder+'\n'
            writesettings()
            if askexport==0:
                lang[130],f"{lang[131]}{exportsavefolder}"
        
        c=1
        path=exportsavefolder+"/"+f'EyePort Data Export {c}.xlsx'
        while os.path.exists(path):
            c=c+1
            path=exportsavefolder+"/"+f'EyePort Data Export {c}.xlsx'
        if c!=1 and overexport==1:
            path=exportsavefolder+"/"+f'EyePort Data Export {c-1}.xlsx'
            d=c-1
        else:
            d=c

        if askexport==1:
            askpath=tk.filedialog.asksaveasfilename(initialdir=exportsavefolder,initialfile=f'EyePort Data Export {d}.xlsx',filetypes=[("Excel File","*.xlsx")])
            if askpath!='':
                if askpath[-5:]!='.xlsx':
                    askpath=askpath+'.xlsx'
                path=askpath
        
        try:
            wb = xls.Workbook()
            ws = wb.active
            ws.title = "Data"
            c1 = ws.cell(row=1, column=1)
            c1.value = "No"
            c2 = ws.cell(row=1, column=2)
            c2.value = "Area of Interest"
            c3 = ws.cell(row=1, column=3)
            c3.value = "Start Time (s)"
            c4 = ws.cell(row=1, column=4)
            c4.value = "End Time (s)"
            c5 = ws.cell(row=1, column=5)
            c5.value = "Duration (s)"
            c6 = ws.cell(row=1, column=6)
            c6.value = "Description"
            c7 = ws.cell(row=1, column=7)
            c7.value = "Gaze Point 2D X"
            c8 = ws.cell(row=1, column=8)
            c8.value = "Gaze Point 2D Y"
            c9 = ws.cell(row=1, column=10)
            c9.value = "No"
            c10 = ws.cell(row=1, column=11)
            c10.value = "Head Orientation"
            c11 = ws.cell(row=1, column=12)
            c11.value = "Start Time (s)"
            c12 = ws.cell(row=1, column=13)
            c12.value = "End Time (s)"
            c13 = ws.cell(row=1, column=14)
            c13.value = "Duration (s)"

            for i in range(len(AOIName)):
                if round(EndSec[i]-StartSec[i],2)<(thit/100)+1:
                    desc='Short'
                elif round(EndSec[i]-StartSec[i],2)<(thit/100)+2:
                    desc='Medium'
                elif round(EndSec[i]-StartSec[i],2)<(thit/100)+5:
                    desc='Long'
                else:
                    desc='Prolonged'
                cell = ws.cell(row=i+2,column=1)
                cell.value = i+1
                cell = ws.cell(row=i+2,column=2)
                cell.value = AOIName[i]
                cell = ws.cell(row=i+2,column=3)
                cell.value = round(StartSec[i],2)
                cell = ws.cell(row=i+2,column=4)
                cell.value = round(EndSec[i],2)
                cell = ws.cell(row=i+2,column=5)
                cell.value = round(EndSec[i]-StartSec[i],2)
                cell = ws.cell(row=i+2,column=6)
                cell.value = desc
                cell = ws.cell(row=i+2,column=7)
                cell.value = GazeX[i]
                cell = ws.cell(row=i+2,column=8)
                cell.value = GazeY[i]

            l=0
            for i in range(len(headtable)):
                cell = ws.cell(row=i+2,column=10)
                l = l + 1
                cell.value = l
                cell = ws.cell(row=i+2,column=11)
                cell.value = headtable[i][0]
                cell = ws.cell(row=i+2,column=12)
                cell.value = round(headtable[i][1],2)
                cell = ws.cell(row=i+2,column=13)
                cell.value = round(headtable[i][2],2)
                cell = ws.cell(row=i+2,column=14)
                cell.value = round(headtable[i][2]-headtable[i][1],2)

            wb.save(path)
            exportstatus = 1
            exportsuccessful()
            print("Exported XLSX File")
        except PermissionError:
            exportstatus = 2
            exportfailed()
            tk.messagebox.showerror(lang[122],lang[132])
            return 0
        except:
            exportstatus = 2
            exportfailed()
            tk.messagebox.showerror(lang[133],lang[134])
            return 0

    def exportsuccessful():
        exportsuccess.grid_forget()
        exportfail.grid_forget()
        export3.grid_forget()
        export4.grid_forget()
        exportsuccess.grid(row=1,column=1,columnspan=2,sticky='nsew')
        export3.grid(row=2,column=1,sticky='se',pady=(25,0),padx=(0,10))
        export4.grid(row=2,column=2,sticky='sw',pady=(25,0),padx=(10,0))
        export3.configure(state='normal')
        export4.configure(state='normal')

    def exportfailed():
        exportsuccess.grid_forget()
        exportfail.grid_forget()
        export3.grid_forget()
        export4.grid_forget()
        exportfail.grid(row=1,column=1,columnspan=2,sticky='nsew')
        export3.grid(row=2,column=1,sticky='se',pady=(25,0),padx=(0,10))
        export4.grid(row=2,column=2,sticky='sw',pady=(25,0),padx=(10,0))
        export3.configure(state='disabled')
        export4.configure(state='disabled')
        

    def export_reset():
        global exportstatus
        exportstatus = 0
        exportclick.grid_forget()
        exportsuccess.grid_forget()
        exportfail.grid_forget()
        export3.grid_forget()
        export4.grid_forget()

    def openexplorer():
        os.system(f'start "" "{exportsavefolder}"')

    def dynafram():
        appnames = AppOpener.give_appnames()
        if 'dynafram' in appnames:
            AppOpener.open('dynaFRAM')
        else:
            ans=tk.messagebox.askyesno(lang[141],lang[142])
            if ans:
                webbrowser.open("https://www.engr.mun.ca/~d.smith/dynafram.html")

    #Navigation and Settings Functions
    def fullscreen():
        if fullscreenmode==1:
            root.state('zoomed')
            root.resizable(False,False)
        else:
            root.state('normal')
            root.resizable(True,True)
            if int(value[3])<101:
                x = (userwinx-1058)//2
                y = (userwiny-620)//2
                root.minsize(1058,620)
                root.geometry(f"1058x620+{x}+{y}")
            elif int(value[3])==110:
                x = (userwinx-1230)//2
                y = (userwiny-720)//2
                root.minsize(1230,720)
                root.geometry(f"1230x720+{x}+{y}")
            else:
                x = (userwinx-1315)//2
                y = (userwiny-770)//2
                root.minsize(1315,770)
                root.geometry(f"1315x770+{x}+{y}")

    def destroyer(relaunch=False):
        global root
        if connected:
            ans = tk.messagebox.askyesno(lang[143],lang[144]) #Connected Warning
        else:
            if relaunch==False:
                ans = tk.messagebox.askyesno(lang[145],lang[146]) #Confirm Quit
            else:
                ans = True
        if ans:
            try:
                plt.close("all")
            except:
                pass
            image1.close()
            image2.close()
            image3.close()
            image4.close()
            image5.close()
            image6.close()
            image7.close()
            image80.close()
            image11.close()
            image12.close()
            image13.close()
            image14.close()
            image15.close()
            image16.close()
            image17.close()
            image18.close()
            image19.close()
            image20.close()
            image21.close()
            image22.close()
            image23.close()
            image24.close()
            image25.close()
            image26.close()
            image27.close()
            image28.close()
            image29.close()
            image291.close()
            image30.close()
            image31.close()
            image32.close()
            image33.close()
            image34.close()
            image35.close()
            image36.close()
            image37.close()
            image38.close()
            image39.close()
            image40.close()
            image41.close()
            image42.close()
            image43.close()
            image44.close()
            image45.close()
            image46.close()
            image47.close()
            image48.close()
            image49.close()
            image50.close()
            image51.close()
            image52.close()
            image53.close()
            image54.close()
            image55.close()
            image56.close()
            image57.close()
            image58.close()
            image59.close()
            image60.close()
            image61.close()
            image62.close()
            image63.close()
            image64.close()
            image65.close()
            image66.close()
            image67.close()
            image68.close()
            image69.close()
            image70.close()
            image71.close()
            image72.close()
            image73.close()
            image74.close()
            image75.close()
            image76.close()
            image77.close()
            image78.close()
            image79.close()
            image80.close()
            image81.close()
            image82.close()
            image83.close()
            image84.close()
            image85.close()
            image86.close()
            image87.close()
            image88.close()
            image89.close()
           
            if relaunch:
                root.destroy()
                root = CTk()
                main()
            else:
                root.destroy()
            _thread.exit()

    def previewtoggle():
        global popenabled
        status = switch_1.get()
        if status == 1:
            popenabled = True
        else:
            popenabled = False
        value[7]=str(status)+"\n"
        writesettings()

    def updatetheme():
        global settingsframe
        chk=theme_var.get()
        if chk==0:
            set_appearance_mode("Light")
            value[2]="Light\n"
        elif chk==1:
            set_appearance_mode("Dark")
            value[2]="Dark\n"
        else:
            set_appearance_mode("System")
            value[2]="System\n"
        try:
            if value[2][:-1]=="Dark":
                styletree.configure("Treeview", background="#333333",fieldbackground="#333333", foreground="white")
            else:
                styletree.configure("Treeview", background="white",fieldbackground="white", foreground="black")
        except:
            pass   
        writesettings()

    def updatescale(new_scaling: str):
        value[3]=new_scaling.replace("%","")+"\n"
        if int(value[3])<101:
            x = (userwinx-1058)//2
            y = (userwiny-620)//2
            root.minsize(1058,620)
            root.geometry(f"1058x620+{x}+{y}")
        elif int(value[3])==110:
            x = (userwinx-1230)//2
            y = (userwiny-720)//2
            root.minsize(1230,720)
            root.geometry(f"1230x720+{x}+{y}")
        elif int(value[3])==120:
            x = (userwinx-1315)//2
            y = (userwiny-770)//2
            root.minsize(1315,770)
            root.geometry(f"1315x770+{x}+{y}")

        if int(value[3][:-1])>=100:
            if combolist[0].cget("size")[0]!=144:
                for i in combolist:
                    i.configure(size=(144,88))
        else:
            if combolist[0].cget("size")[0]!=180:
                for i in combolist:
                    i.configure(size=(180,110))
            
        new_scaling_float = int(new_scaling.replace("%","")) / 100
        set_widget_scaling(new_scaling_float)
        writesettings()

    def updateasense(a: str):
        global asense
        if a=='Extremely Low':
            asense=0.50
        elif a=='Very Low':
            asense=0.60
        elif a=='Low':
            asense=0.70
        elif a=='Normal':
            asense=0.75
        elif a=='High':
            asense=0.80
        else:
            asense=0.85
        value[12]=a+"\n"
        writesettings()

    def updateuseai(override=False):
        global useai
        if not override:
            useai = useai_var.get()
        else:
            radio_button_8.select()
        value[13]=str(useai)+"\n"
        writesettings()

    def updatelang():
        global language_mode,lang
        language_mode = lang_var.get()
        value[24]=str(language_mode)+"\n"
        writesettings()
        if language_mode==0:
            lang = English
            ans = tk.messagebox.askyesno("Relaunch EyePort","EyePort needs to relaunch to apply your selected language. Relaunch Now?")
        elif language_mode==1:
            lang = French
            ans = tk.messagebox.askyesno("Relancer EyePort","La traduction peut être incorrecte à certains endroits. EyePort doit être relancé pour appliquer la langue sélectionnée. Relancer maintenant?")
        else:
            lang = Dutch
            ans = tk.messagebox.askyesno("EyePort opnieuw starten", "De vertaling kan op sommige plaatsen onjuist zijn. EyePort moet opnieuw starten om de door u geselecteerde taal toe te passen. Nu opnieuw starten?")
        if ans:
            destroyer(ans)

    def updatetimehit(a):
        global thit
        thit = a
        if thit%50==0:
            text70.configure(text=str(thit/100)+'0 s')
        else:
            text70.configure(text=str(thit/100)+' s')
        value[17] = str(int(a))+"\n"
        writesettings()

    def updateaviolate(a):
        global aviolate
        aviolate = int(a)
        text79.configure(text=str(aviolate)+' s')
        value[5] = str(int(a))+"\n"
        writesettings()

    def updateadead(a):
        global adead
        adead = int(a)
        text82.configure(text=str(adead)+' s')
        value[6] = str(int(a))+"\n"
        writesettings()

    def updateatolbox(a):
        global atolbox
        atolbox = int(a)
        text83.configure(text=str(atolbox)+' px')
        value[8] = str(int(a))+"\n"
        updatetolpic()
        writesettings()

    def updatetolpic():
        if int(value[8][:-1])==50 or int(value[8][:-1])==100:
            combo11.configure(image=imgtk43)
        elif int(value[8][:-1])==150 or int(value[8][:-1])==200:
            combo11.configure(image=imgtk44)
        elif int(value[8][:-1])==250 or int(value[8][:-1])==300:
            combo11.configure(image=imgtk45)
        elif int(value[8][:-1])==350 or int(value[8][:-1])==400:
            combo11.configure(image=imgtk46)
        elif int(value[8][:-1])==450 or int(value[8][:-1])==500:
            combo11.configure(image=imgtk47)
        elif int(value[8][:-1])==550 or int(value[8][:-1])==600:
            combo11.configure(image=imgtk48)
        elif int(value[8][:-1])==650 or int(value[8][:-1])==700:
            combo11.configure(image=imgtk49)
        elif int(value[8][:-1])==750 or int(value[8][:-1])==800:
            combo11.configure(image=imgtk50)
        else:
            combo11.configure(image=imgtk51)

    def updateheadsen(a):
        global headsen
        headsen = a
        g = str(headsen)
        if len(g)>3:
            g = g[:2]
        else:
            g = g[:1]
        text76.configure(text=g+' deg/s')
        value[4] = str(int(a))+"\n"
        updateheadsenpic()
        writesettings()

    def updateheadsenpic():
        if int(value[4][:-1])<20:
            combo12.configure(image=imgtk55)
        elif int(value[4][:-1])>=20 and int(value[4][:-1])<40:
            combo12.configure(image=imgtk52)
        elif int(value[4][:-1])>=40 and int(value[4][:-1])<60:
            combo12.configure(image=imgtk53)
        else:
            combo12.configure(image=imgtk54)

    def rememberload():
        global rememberloadbool
        status = check_2.get()
        if status==1:
            rememberloadbool=True
        else:
            rememberloadbool=False
        value[14]=str(status)+"\n"
        writesettings()

    def updateheadmode():
        global headmode
        headmode=headmode_var.get()
        value[20]=str(headmode)+"\n"
        writesettings()

    def frameone_forget():
        text80.pack_forget()
        heading1.pack_forget()
        heading2.pack_forget()
        recorddata.pack_forget()
        analyzedata.pack_forget()
        viewdata.pack_forget()
        exportdata.pack_forget()
        settingsbutton.pack_forget()
        text1.pack_forget()
        quitbutton.pack_forget()
        usermanual2.pack_forget()

    def updatesidebar():
        global compact
        compact = check_4.get()
        frameone_forget()
        if compact==0:
            recorddata.configure(text=lang[7],height=bheight,width=bwidth,corner_radius=5,compound="left",fg_color=("#3b8ed0","#1f6aa5"))
            analyzedata.configure(text=lang[5],height=bheight,width=bwidth,corner_radius=5,compound="left",fg_color=("#3b8ed0","#1f6aa5"))
            viewdata.configure(text=lang[6],height=bheight,width=bwidth,corner_radius=5,compound="left",fg_color=("#3b8ed0","#1f6aa5"))
            exportdata.configure(text=lang[4],height=bheight,width=bwidth,corner_radius=5,compound="left",fg_color=("#3b8ed0","#1f6aa5"))
            settingsbutton.configure(text=lang[8],height=bheight,width=bwidth,corner_radius=5,compound="left",fg_color=("#3b8ed0","#1f6aa5"))
            quitbutton.configure(text=lang[3],height=bheight,width=bwidth,corner_radius=5,compound="left")
            usermanual2.configure(text=lang[14],height=bheight,width=bwidth,corner_radius=5,compound="left",fg_color=("#3b8ed0","#1f6aa5"))
            text80.pack(side=tk.TOP,pady=(20,0))
            heading1.pack(side=tk.TOP,pady=(10,0))
            heading2.pack(side=tk.TOP,pady=(0,10))
            recorddata.pack(side=tk.TOP,pady=10,padx=20)
            analyzedata.pack(side=tk.TOP,pady=10,padx=20)
            viewdata.pack(side=tk.TOP,pady=10,padx=20)
            exportdata.pack(side=tk.TOP,pady=10,padx=20)
            settingsbutton.pack(side=tk.TOP,pady=10,padx=20)
            text1.pack(side=tk.BOTTOM,pady=10,padx=20)
            quitbutton.pack(side=tk.BOTTOM,pady=(10,0),padx=20)
            usermanual2.pack(side=tk.BOTTOM,pady=10,padx=20)
        else:
            recorddata.configure(text="",height=bsheight,width=bswidth,corner_radius=0,compound="bottom")
            recorddata.pack(side=tk.TOP,pady=0,padx=0)
            analyzedata.configure(text="",height=bsheight,width=bswidth,corner_radius=0,compound="bottom")
            analyzedata.pack(side=tk.TOP,pady=0,padx=0)
            if viewdata.cget("state")=="disabled":
                viewdata.configure(text="",height=bsheight,width=bswidth,corner_radius=0,compound="bottom",fg_color=("#cfcfcf","#333333"))
            else:
                viewdata.configure(text="",height=bsheight,width=bswidth,corner_radius=0,compound="bottom",fg_color=("#3b8ed0","#1f6aa5"))
            viewdata.pack(side=tk.TOP,pady=0,padx=0)
            exportdata.configure(text="",height=bsheight,width=bswidth,corner_radius=0,compound="bottom")
            exportdata.pack(side=tk.TOP,pady=0,padx=0)
            settingsbutton.configure(text="",height=bsheight,width=bswidth,corner_radius=0,compound="bottom",fg_color=("#36719f","#144870"))
            settingsbutton.pack(side=tk.TOP,pady=0,padx=0)
            quitbutton.configure(text="",height=bsheight,width=bswidth,corner_radius=0,compound="bottom")
            quitbutton.pack(side=tk.BOTTOM,pady=(0,0),padx=0)
            usermanual2.configure(text="",height=bsheight,width=bswidth,corner_radius=0,compound="bottom")
            usermanual2.pack(side=tk.BOTTOM,pady=(0,0),padx=0)
        value[11] = str(compact)+"\n"
        writesettings()

    def overexporter():
        global overexport
        overexport=check_5.get()
        value[21]=str(overexport)+"\n"
        writesettings()

    def askexporter():
        global askexport
        askexport=check_6.get()
        value[22]=str(askexport)+"\n"
        writesettings()

    def usemanual():
        global use_manual
        use_manual=check_7.get()
        value[23]=str(use_manual)+"\n"
        writesettings()

    def usefullscreen():
        global fullscreenmode
        fullscreenmode = check_3.get()
        fullscreen()
        value[18]=str(fullscreenmode)+"\n"
        writesettings()

    #View Page
    def resizeview(e):
        global viewx,viewy
        viewx=e.width
        viewy=e.height

    def update_duration(e):
        global raw
        raw = player.video_info()["duration"]
        viewslider.configure(to=raw)
        min = int(raw//60)
        sec = floor(raw)%60
        viewend.configure(text=f'{min}:{sec}')

    def update_scale(e):
        global viewadvancer,viewadvancer1
        curr = player.current_duration()
        if overlay==1:
            index = int(ceil((curr/raw)*len(Timestamps)))
            try:
                placey=(0.804*(viewy*GazeYmin[index]))-116.236
                placex=(0.796*(viewx*GazeXmin[index]))-69.72
                if curr>headtable[viewadvancer1][2]:
                    viewadvancer1 = viewadvancer1+1
                    viewtext.configure(text=f" {AOIName[viewadvancer]} | {headtable[viewadvancer1][0]} ")
                if curr>EndSec[viewadvancer]:
                    viewadvancer = viewadvancer+1
                    if exp:
                        img = Image.fromarray(picarray[viewadvancer])
                        tk_image = CTkImage(light_image=img, size=(192,108))
                        viewtext.configure(text=f" {AOIName[viewadvancer]} | {headtable[viewadvancer1][0]} ",image=tk_image)
                    else:
                        viewtext.configure(text=f" {AOIName[viewadvancer]} | {headtable[viewadvancer1][0]} ")
                viewtext.place(y=placey,x=placex)
            except:
                pass
        viewslider.set(curr)

    def video_ended(e):
        global viewadvancer
        playpausebutton.configure(image=imgtk151)
        viewtext.place_forget()
        viewadvancer = 0
        viewslider.set(0)

    def skip(value,event=None):
        player.seek(int(viewslider.get())+value)
        viewslider.set(viewslider.get()+value)
        if overlay==1:
            checkview()

    def checkview():
        global viewadvancer
        viewadvancer = 0
        while viewslider.get()>EndSec[viewadvancer]:
            viewadvancer = viewadvancer + 1
        viewtext.configure(text=f" {AOIName[viewadvancer]} ")

    def seek(value):
        if overlay==1:
            checkview()
        player.seek(int(value))

    def leftbutton(event=None):
        skip(-1)

    def rightbutton(event=None):
        skip(1)

    def playvideo(event=None):
        if player.is_paused():
            player.play()
            playpausebutton.configure(image=imgtk28)
        else:
            player.pause()
            playpausebutton.configure(image=imgtk151)

    def toggleoverlay():
        global overlay
        if overlay==0:
            overlay=1
        else:
            overlay=0
            viewtext.place_forget()
        value[9]=str(overlay)+"\n"
        writesettings()

    #Welcome Page
    def showagain():
        val = showagain.get()
        if val==1:
            check_8.deselect()
        else:
            check_8.select()
        value[10]=str(val)+'\n'
        writesettings()

    def updatewelcome():
        val = check_8.get()
        if val==1:
            welcome = False
        else:
            welcome = True
        value[10]=str(val)+'\n'
        writesettings()

    #User Manual
    def showusermanual():
        if language_mode==0:
            path = ".\EyePort User Manual.pdf"
        elif language_mode==1:
            path = ".\EyePort User Manual FR.pdf"
        else:
            path = ".\EyePort User Manual DU.pdf"
        if os.path.exists(path):
            subprocess.Popen([path], shell=True)
        else:
            tk.messagebox.showerror(lang[135],lang[136])      

    #Page Loading
    def analyzepage():
        usermanual2.pack_forget()
        if compact==0:
            usermanual2.pack(side=tk.BOTTOM,pady=10,padx=20)
        else:
            usermanual2.pack(side=tk.BOTTOM,pady=0,padx=0)
            for i in framebuttons:
                i.configure(fg_color=("#3b8ed0","#1f6aa5"))
            analyzedata.configure(fg_color=("#36719f","#144870"))
        recordframe.pack_forget()
        exportframe.pack_forget()
        settingsframe.pack_forget()
        viewframe.pack_forget()
        welcomeframe.destroy()
        videoframe.unbind('<Configure>')
        expframe.unbind('<Configure>')
        viewframe1.unbind('<Configure>')
        root.unbind('<space>')
        root.unbind('<Left>')
        root.unbind('<Right>')
        analyzeframe.pack(fill=tk.BOTH,side=tk.LEFT,expand=True)
        analyzeframe.grid_columnconfigure((0,1,2),weight=1)
        
        text2.grid(row=0,column=0,sticky='nw',padx=(20,0),pady=12)
        
        text12.grid(row=1,column=0,columnspan=2,sticky='w',padx=(20,0),pady=(5,0))
        text13.grid(row=2,column=0,columnspan=2,sticky='w',padx=(30,0))
        loadfile.grid(row=1,column=2,rowspan=2,padx=(0,20),pady=(5,0),sticky='e')
        
        text42.grid(row=5,column=0,columnspan=2,sticky='w',padx=(20,0),pady=(5,0))
        text43.grid(row=6,column=0,columnspan=2,sticky='w',padx=(30,0))
        analyze.grid(row=5,column=2,rowspan=2,padx=(0,20),pady=(5,0),sticky='e')

        progressframe.grid(row=7,column=0,columnspan=3,sticky='nsew',padx=20,pady=20)
        progressframe.grid_columnconfigure((0,1),weight=1)
        text84.grid(row=0,column=0,columnspan=2,padx=(20,0),pady=(10,0),sticky='w')
        text85.grid(row=1,column=0,padx=(20,0),pady=(2,0),sticky='w')
        text86.grid(row=1,column=1,padx=(0,20),pady=(2,0),sticky='e')
        progressbar.grid(row=2,column=0,columnspan=2,padx=(20,20),pady=(2,15),sticky='nsew')

    def viewpage():
        usermanual2.pack_forget()
        if compact==0:
            usermanual2.pack(side=tk.BOTTOM,pady=10,padx=20)
        else:
            usermanual2.pack(side=tk.BOTTOM,pady=0,padx=0)
            for i in framebuttons:
                i.configure(fg_color=("#3b8ed0","#1f6aa5"))
            viewdata.configure(fg_color=("#36719f","#144870"))
        viewframe1.bind('<Configure>',resizeview)
        expframe.unbind('<Configure>')
        videoframe.unbind('<Configure>')
        root.bind('<space>',playvideo)
        root.bind('<Left>',leftbutton)
        root.bind('<Right>',rightbutton)
        recordframe.pack_forget()
        analyzeframe.pack_forget()
        settingsframe.pack_forget()
        exportframe.pack_forget()
        welcomeframe.destroy()
        viewframe.pack(fill=tk.BOTH,side=tk.LEFT,expand=True)
        viewframe.columnconfigure((0,1),weight=1)
        viewframe.rowconfigure(1,weight=1)
        check_1.grid(row=0,column=1,sticky='e',padx=(0,20),pady=(12,0))
        text36.grid(row=0,column=0,sticky='nw',padx=(20,0),pady=(12,0))
        viewframe1.grid(row=1,column=0,columnspan=2,sticky='nsew',padx=20,pady=20)
        viewframe1.rowconfigure(0,weight=1)
        viewframe1.columnconfigure((0,4),weight=1)
        viewstart.grid(row=1,column=0,sticky='w',padx=15)
        viewend.grid(row=1,column=4,sticky='e',padx=15)
        playpausebutton.grid(row=1,column=2,pady=10,padx=5)
        forwardbutton.grid(row=1,column=1,pady=10)
        backbutton.grid(row=1,column=3,pady=10)
        viewslider.grid(row=2,column=0,columnspan=5,padx=10,pady=(0,10),sticky='nsew')

    def exportpage():
        usermanual2.pack_forget()
        if compact==0:
            usermanual2.pack(side=tk.BOTTOM,pady=10,padx=20)
        else:
            usermanual2.pack(side=tk.BOTTOM,pady=0,padx=0)
            for i in framebuttons:
                i.configure(fg_color=("#3b8ed0","#1f6aa5"))
            exportdata.configure(fg_color=("#36719f","#144870"))
        expframe.unbind('<Configure>')
        viewframe1.unbind('<Configure>')
        root.unbind('<space>')
        root.unbind('<Left>')
        root.unbind('<Right>')
        expframe.rowconfigure((0,3),weight=1)
        expframe.columnconfigure((0,3),weight=1)
        if exportsavefolder!='Expsave':
            text54.configure(text=f"{lang[105]}{exportsavefolder}")
            exportsave.configure(text=lang[89],fg_color='Green',hover_color="#006400") #Auto Selected
        recordframe.pack_forget()
        analyzeframe.pack_forget()
        settingsframe.pack_forget()
        viewframe.pack_forget()
        welcomeframe.destroy()
        exportframe.pack(fill=tk.BOTH,side=tk.LEFT,expand=True)
        exportframe.columnconfigure(0,weight=1)
        exportframe.rowconfigure(4,weight=1)
        text4.grid(row=0,column=0,sticky='nw',padx=(20,0),pady=12)

        text53.grid(row=1,column=0,sticky='w',padx=(20,0),pady=(5,0))
        text54.grid(row=2,column=0,sticky='w',padx=(30,0))
        exportsave.grid(row=1,column=1,rowspan=2,padx=(0,20),pady=(5,0),sticky='e')
        
        expframe1.grid(row=3,column=0,columnspan=2,padx=20,pady=(20,0))
        export1.grid(row=0,column=0,padx=(0,10),sticky='e')
        export2.grid(row=0,column=1,padx=(0,0),sticky='e')
        export5.grid(row=0,column=2,padx=(10,0),sticky='w')

        expframe.grid(row=4,column=0,columnspan=2,sticky='nsew',padx=20,pady=20)

    def recordpage():
        usermanual2.pack_forget()
        if compact==0:
            usermanual2.pack(side=tk.BOTTOM,pady=10,padx=20)
        else:
            usermanual2.pack(side=tk.BOTTOM,pady=0,padx=0)
            for i in framebuttons:
                i.configure(fg_color=("#3b8ed0","#1f6aa5"))
            recorddata.configure(fg_color=("#36719f","#144870"))
        videoframe.bind('<Configure>',resizerecord)
        expframe.unbind('<Configure>')
        viewframe1.unbind('<Configure>')
        root.unbind('<space>')
        root.unbind('<Left>')
        root.unbind('<Right>')
        analyzeframe.pack_forget()
        exportframe.pack_forget()
        settingsframe.pack_forget()
        viewframe.pack_forget()
        welcomeframe.destroy()
        recordframe.pack(fill=tk.BOTH,side=tk.LEFT,expand=True)
        recordframe.columnconfigure(3,weight=1)
        recordframe.rowconfigure(2,weight=1)
        text3.grid(row=0,column=0,columnspan=2,sticky='nw',padx=(20,0),pady=(12,0))
        switch_1.grid(row=0,column=3,sticky='e',padx=(0,20),pady=(15,0)) #Pop Live View
        
        connectbutton.grid(row=1,column=0,padx=(20,0),pady=20,sticky='w')
        calibratebutton.grid(row=1,column=1,padx=(10,0),pady=20,sticky='w')
        previewbutton.grid(row=1,column=2,padx=(10,0),pady=20,sticky='w')
        recordbutton.grid(row=1,column=3,padx=(10,0),pady=20,sticky='w')
        
        videoframe.grid(row=2,column=0,columnspan=4,sticky='nsew',padx=20,pady=(0,20))
        videoframe.grid_columnconfigure(0,weight=1)
        videoframe.grid_rowconfigure(1,weight=1)
        text20.grid(row=0,column=0,padx=(20,0),pady=(10,0),sticky='w')
        if connected:
            if stopper:
                previewavailfun()
            else:
                pass
        else:
            connectpreviewfun()

    def settingspage():
        usermanual2.pack_forget()
        if compact==0:
            usermanual2.pack(side=tk.BOTTOM,pady=10,padx=20)
        else:
            usermanual2.pack(side=tk.BOTTOM,pady=0,padx=0)
            for i in framebuttons:
                i.configure(fg_color=("#3b8ed0","#1f6aa5"))
            settingsbutton.configure(fg_color=("#36719f","#144870"))
        videoframe.unbind('<Configure>')
        expframe.unbind('<Configure>')
        viewframe1.unbind('<Configure>')
        root.unbind('<space>')
        root.unbind('<Left>')
        root.unbind('<Right>')
        analyzeframe.pack_forget()
        exportframe.pack_forget()
        viewframe.pack_forget()
        recordframe.pack_forget()
        welcomeframe.destroy()
        settingsframe.pack(fill=tk.BOTH,side=tk.LEFT,expand=True)
        text5.grid(row=0,column=0,sticky='nw',padx=(20,0),pady=(12,0)) #Settings
        settingsframe.grid_columnconfigure(0,weight=1)
        settingstab.grid(row=1,column=0,sticky='nsew',padx=20,pady=12)

        text71.grid(row=0,column=0,sticky='nw',padx=20,pady=(5,0)) #Language
        frame3.grid(row=1,column=0,sticky='nw',padx=20,pady=(2,0)) #Options
        radio_button_4.pack(side=tk.TOP,ipadx=3,ipady=3,padx=3,pady=3)
        radio_button_5.pack(side=tk.TOP,ipadx=3,ipady=3,padx=3,pady=3)
        radio_button_6.pack(side=tk.TOP,ipadx=3,ipady=3,padx=3,pady=3)
        text7.grid(row=2,column=0,sticky='nw',padx=20,pady=(10,0)) #Theme
        frame2.grid(row=3,column=0,sticky='nw',padx=20,pady=(2,0)) #Options
        combo1.grid(row=0,column=0,padx=(10,0),pady=(10,0))
        combo2.grid(row=0,column=1,padx=(10,0),pady=(10,0))
        combo3.grid(row=0,column=2,padx=10,pady=(10,0))
        radio_button_1.grid(row=1,column=0,sticky='w',padx=10,pady=10)
        radio_button_2.grid(row=1,column=1,sticky='w',padx=10,pady=10)
        radio_button_3.grid(row=1,column=2,sticky='w',padx=10,pady=10)
        text8.grid(row=4,column=0,sticky='nw',padx=20,pady=(10,0)) #UI Scaling
        scaling_optionmenu.grid(row=5,column=0,sticky='nw',padx=20,pady=(2,0)) #ComboBox
        check_3.grid(row=6,column=0,sticky='nw',padx=20,pady=(15,0)) #Fullscreen
        check_4.grid(row=7,column=0,sticky='nw',padx=20,pady=(7,0)) #Collapse
        check_8.grid(row=8,column=0,sticky='nw',padx=20,pady=(7,0)) #Welcome
        reset.grid(row=9,column=0,sticky='nw',padx=20,pady=(15,15)) #Reset Settings
        
        settingstab.tab(lang[166]).columnconfigure(1,weight=1)
        text68.grid(row=0,column=0,sticky='w',padx=20,pady=(15,0)) #Fixation
        slider_2.grid(row=0,column=1,sticky='e',padx=(0,0),pady=(15,0))
        text70.grid(row=0,column=2,sticky='w',padx=(10,0),pady=(15,0)) #s Text
        text68help.grid(row=0,column=3,sticky='w',padx=(0,20),pady=(15,0))
        text77.grid(row=1,column=0,sticky='w',padx=20,pady=(10,0)) #Radar
        slider_4.grid(row=1,column=1,sticky='e',padx=(0,0),pady=(10,0)) 
        text79.grid(row=1,column=2,sticky='w',padx=(10,0),pady=(10,0)) #s Text
        text77help.grid(row=1,column=3,sticky='w',padx=(0,20),pady=(10,0))
        text81.grid(row=2,column=0,sticky='w',padx=20,pady=(10,0)) #Dead Man
        slider_1.grid(row=2,column=1,sticky='e',padx=(0,0),pady=(10,0)) 
        text82.grid(row=2,column=2,sticky='w',padx=(10,0),pady=(10,0)) #s Text
        text81help.grid(row=2,column=3,sticky='w',padx=(0,20),pady=(10,0))
        text65.grid(row=3,column=0,columnspan=3,sticky='nw',padx=20,pady=(15,0)) #Object Detection
        if not imageai_avail:
            text66.grid(row=4,column=0,columnspan=4,sticky='nw',padx=20,pady=(2,0))
            combo4.configure(image=imgtk25)
            combo5.configure(image=imgtk24)
            combo6.configure(image=imgtk40)
            combo7.configure(image=imgtk29)
        frame6.grid(row=5,column=0,columnspan=4,sticky='nw',padx=20,pady=(2,0)) 
        combo4.grid(row=0,column=0,padx=(10,0),pady=(10,0))
        combo5.grid(row=0,column=1,padx=(10,0),pady=(10,0))
        combo6.grid(row=0,column=2,padx=(10,0),pady=(10,0))
        combo7.grid(row=0,column=3,padx=10,pady=(10,0))
        radio_button_8.grid(row=1,column=0,sticky='w',padx=10,pady=10)
        radio_button_9.grid(row=1,column=1,sticky='w',padx=10,pady=10)
        radio_button_10.grid(row=1,column=2,sticky='w',padx=10,pady=10)
        radio_button_11.grid(row=1,column=3,sticky='w',padx=10,pady=10)
        
        text72.grid(row=6,column=0,columnspan=3,sticky='nw',padx=20,pady=(15,0)) #Head Orientation
        text72help.grid(row=6,column=3,sticky='w',padx=(0,20),pady=(15,0))
        frame4.grid(row=7,column=0,columnspan=4,sticky='nw',padx=20,pady=(2,0)) 
        combo8.grid(row=0,column=0,padx=(10,0),pady=(10,0))
        combo9.grid(row=0,column=1,padx=(10,0),pady=(10,0))
        combo10.grid(row=0,column=2,padx=10,pady=(10,0))
        radio_button_12.grid(row=1,column=0,sticky='w',padx=10,pady=10)
        radio_button_13.grid(row=1,column=1,sticky='w',padx=10,pady=10)
        radio_button_14.grid(row=1,column=2,sticky='w',padx=10,pady=10)
        check_2.grid(row=9,column=0,sticky='nw',padx=20,pady=(15,0)) #Remember
        check_5.grid(row=10,column=0,sticky='nw',padx=20,pady=(7,0)) #Overwrite Old Exports
        check_6.grid(row=11,column=0,sticky='nw',padx=20,pady=(7,0)) #Ask for File Names
        nogozone.grid(row=12,column=0,sticky='nw',padx=20,pady=(15,15)) #No Go Zone

        settingstab.tab(lang[167]).columnconfigure(1,weight=1)
        check_7.grid(row=0,column=0,columnspan=4,sticky='nw',padx=(20,0),pady=(10,0)) #Manual Mode
        text47.grid(row=1,column=0,sticky='w',padx=(20,0),pady=(15,0)) #Detection Sensitivity
        scaling_optionmenu2.grid(row=1,column=1,columnspan=2,sticky='e',padx=(10,0),pady=(10,0))
        text47help.grid(row=1,column=3,sticky='w',padx=(0,20),pady=(10,0))
        text78.grid(row=2,column=0,sticky='w',padx=(20,0),pady=(15,0)) #Tolerance Box Size
        slider_5.grid(row=2,column=1,sticky='e',padx=(0,0),pady=(10,0))
        text83.grid(row=2,column=2,sticky='w',padx=(10,0),pady=(10,0)) #px
        text78help.grid(row=2,column=3,sticky='w',padx=(0,20),pady=(10,0))
        combo11.grid(row=3,column=0,columnspan=4,sticky='w',padx=(20,0),pady=(15,0))
        text74.grid(row=4,column=0,sticky='w',padx=(20,0),pady=(10,0)) #Head Sensitivity
        slider_3.grid(row=4,column=1,sticky='e',padx=(0,0),pady=(10,0))
        text76.grid(row=4,column=2,sticky='w',padx=(10,0),pady=(10,0)) #deg/s
        text74help.grid(row=4,column=3,sticky='w',padx=(0,20),pady=(10,0))
        combo12.grid(row=5,column=0,columnspan=4,sticky='w',padx=(20,0),pady=(15,15))

    def welcomepage():
        welcomeframe.pack(fill=tk.BOTH,side=tk.LEFT,expand=True)
        welcomeframe.columnconfigure((0,1),weight=1)
        welcomeframe.rowconfigure((0,6),weight=1)
        text37.grid(row=1,column=0,columnspan=2,sticky='nsew')
        text38.grid(row=2,column=0,columnspan=2,sticky='n')
        text39.grid(row=3,column=0,columnspan=2,sticky='n')
        gorecbutton.grid(row=4,column=0,pady=(20,0),padx=5,sticky='e')
        usermanual.grid(row=4,column=1,pady=(20,0),padx=5,sticky='w')
        showagain.grid(row=5,column=0,columnspan=2,pady=(20,50),sticky='n')

    #Getting settings
    if not os.path.exists("Settings.txt"):
        createsettings()

    #For other functions to write to the settings file
    def writesettings():
        f=open("Settings.txt","w")
        f.writelines(value)
        f.close()

    #Normal Reset from app
    def resetsettings2():
        global popenabled,settings,fullscreenmode
        settings=[]
        ans=tk.messagebox.askyesno(lang[139],lang[140])
        if ans:
            settings=[]
            radio_button_1.select()
            scaling_optionmenu.set("100%")
            switch_1.deselect()
            slider_2.set(200)
            slider_3.set(20)
            slider_4.set(30)
            slider_5.set(300)
            text82.configure(text='40 s')
            text70.configure(text='2.00 s')
            text76.configure(text='20 deg/s')
            text79.configure(text='30 s')
            text83.configure(text='300 px')
            combo11.configure(image=imgtk45)
            combo12.configure(image=imgtk55)
            radio_button_8.select()
            radio_button_13.select()
            check_2.select()
            check_4.deselect()
            if compact==1:
                updatesidebar()
            check_3.deselect()
            check_5.deselect()
            check_6.deselect()
            check_7.deselect()
            check_8.select()
            fullscreenmode = False
            scaling_optionmenu2.set("Normal")
            popenabled = False
            text54.configure(text=lang[47]) #Choose Export Folder Desc
            exportsave.configure(text=lang[55],fg_color=("#3b8ed0","#1f6aa5"),hover_color=("#36719f","#144870")) #Export Folder Button
            createsettings()
            applysettings()

    #In case settings files was corrupted
    def resetsettings():
        global settings
        settings=[]
        tk.messagebox.showerror("Settings File Corruption","The Settings File looks corrupted, which may have occurred due to tampering with the software files. Unable to apply your settings. EyePort will start in Default Settings.")
        createsettings()
        applysettings()

    #Boot and apply settings
    def applysettings():
        global value,themeframe,welcome,rememberloadbool,headsen,thit,useai,fullscreenmode,exportsavefolder,overexport,headmode,askexport,use_manual,overlay,language_mode,aviolate,adead,atolbox,compact
        f=open("Settings.txt","a+")
        f.seek(0)
        value=f.readlines()
        check=[['EyePort Settings File - Any corruption may lead to configuration loss'],['Made by Akash Samanta'],['Light','Dark','System'],['80','90','100','110','120'],[str(j) for j in range(5,81,5)],[str(j) for j in range(5,121,5)],[str(j) for j in range(10,121,10)],['0','1'],[str(j) for j in range(50,1001,50)],['0','1'],['0','1'],['0','1'],["Extremely Low","Very Low", "Low", "Normal", "High","Very High"],['0','1','2','3'],['0','1'],[],[],[str(j) for j in range(25,501,25)],['0','1'],[],['1','2','3'],['0','1'],['0','1'],['0','1'],['0','1','2']]
        #print(len(value),len(check)) #For Debug Only
        for i in range(len(check)):
            try:
                if check[i]!=[]:
                    if value[i][:-1] not in check[i]:
                        #print(value[i][:-1],check[i]) #For Debug Only
                        resetsettings()
            except:
                #print("FAIL") #For Debug Only
                resetsettings()
        if value[2][:-1]=="Light":
            set_appearance_mode("Light")
        elif value[2][:-1]=="Dark":
            set_appearance_mode("Dark")
        else:
            set_appearance_mode("System")
        set_widget_scaling(int(value[3][:-1])/100) #Scaling
        headsen = int(value[4][:-1]) #Head Sensitivity
        aviolate = int(value[5][:-1]) #Violation
        adead = int(value[6][:-1]) #Dead Man
        #value[7] Fullscreen Mode
        atolbox = int(value[8][:-1]) #Tolerance Box
        overlay = int(value[9][:-1])
        if value[10][:-1]=="1": #Welcome Screen
            welcome=False
        else:
            welcome=True
        compact = int(value[11][:-1])
        updateasense(value[12][:-1]) #Detection Sensitivity
        useai=int(value[13][:-1]) #Use AI Object Detection
        settings.append(int(value[14][:-1])) #Remember Last Loaded Files
        if int(value[14][:-1])==1:
            rememberloadbool=True
        else:
            rememberloadbool=False
        #value[15] LoadJSON
        #value[16] empty
        thit = int(value[17][:-1])
        fullscreenmode = int(value[18][:-1])
        exportsavefolder=value[19][:-1]
        if os.path.exists(exportsavefolder)==False:
            exportsavefolder='Expsave'
        headmode = int(value[20][:-1])
        overexport = int(value[21][:-1])
        askexport = int(value[22][:-1])
        use_manual = int(value[23][:-1])
        language_mode = int(value[24][:-1])
        #print(settings) #For Debug Only
        f.close()

    #Making the Application
    root.title("EyePort")
    root.protocol("WM_DELETE_WINDOW", destroyer)
    root.iconbitmap(r".\Resources\Logo2.ico")
    applysettings()
    userwinx = root.winfo_screenwidth()
    userwiny = root.winfo_screenheight()

    #Language
    English = ["EyePort","Version ","Made by Akash Samanta","Close EyePort","Export Data","Analyze Data","Playback Data","Record Data","Settings","Analyze Data","Record Data","Export Data","Welcome to EyePort","To get started, go to the \"Record Data\" Page, or if you are new, try reading the User Manual.",
            "User Manual","Do not show this page again on startup","Use Overlays","Analysis Box Size","Language","Theme","Light","Dark","System","UI Scaling","Dead Man Switch Delay","Fullscreen Mode","Pop Live View","Reset All Settings","Define No Go Zones","Define the square side length in pixels. This cropped image will be used for detecting Unique Areas of Interest.","Fixation Time",
            "Choose how long to focus on an object before registering it as an Area of Interest.","Detection Sensitivity","Choose how sensitive the analyzing algorithm should be to detect Unique Areas of Interest.","Automatic Object Detection","Choose if you would like to use Automatic Object Detection.","Disabled","General","Ships and Icebergs","VISTA Diesel Engine",
            "Remember Last Project","Compact Mode","Choose the time delay needed to activate the Dead Man Switch after no response from participant.","Overwrite Old Exports","Ask for File Names before Exporting","Manual Mode","Select Output Folder","Choose where you would like to export your analyzed files.","Collecting Glass Data","Checking Gaze Overlay Data","Clear Page","Dead Man Switch",
            "Detections","Show Welcome Screen","Analyzing","Export Folder","Export CSV","Export XMFV","Export XLSX","Show in Explorer","Open DynaFRAM","Load SD Card Folder","Load your desired project folder from the SD Card.","Start Analysis","Click \"Analyze Now\" to start analyzing. Analysis will take longer if Object Detection is used.",
            "Select Folder","Analyze Now","Head Orientation Times","Head Up Time","Head Level Time","Head Down Time","Edit Names","No-Go Zone Violations","Dead Man Switch: Not Active","Show Graphs","Radar Violations","Calculating Areas of Interests","Calculating Head Orientation Data","Connect Glasses","Synchronizing Timestamps","Getting Frame Data","Calibrate Glasses","Calculating Unique Areas of Interests","Calculating No-Go Zone Violations",
            "Record Now","TOBII Glasses Live View","Not Available","Disconnect Glasses","Start Live View","Auto Selected","Connection Success","Stop Live View","Loading","Stop Recording","Error Starting Recording","EyePort is unable to communicate with the TOBII Glasses. Perhaps check the connection again?","Error Stopping Recording","EyePort is unable to stop the recording. This could be due to a poor connection with the glasses. Your recording will continue until the glasses runs out of battery or SD Card space. Press the power button for 4 seconds for a controlled shut down.",
            "Start Calibration","Please ask the participant to look at the calibration card and then dismiss this prompt to start calibration.","Calibration Failed","Calibration Failure","Please ask the participant to keep looking at the card during the calibration process. There weren't enough valid samples to complete the calibration. Please try again.","Calibrate Again","Project Folder Loaded","Using ","Please load the project folder to start analysis.","Save Changes","Unique Area of Interests: ","Area of Interests: ","Dead Man Switch: Active",
            "Selected","Connection Failed","EyePort is unable to connect with the TOBII Glasses.\nPlease check the connection and try again.","Recording Still in Progress","You have a recording in progress. Please stop the recording and then disconnect the glasses.","Invalid Folder","EyePort was unable to locate glass data in the selected folder. Please select the correct folder and try again.","Failed to assign Unique Names","You cannot assign identical names to Unique Areas of Interest. Use Manual Mode to override this behaviour.","Problem Loading Files","EyePort was unable to open the files. The files may have been moved or deleted. Please check the loaded files in file explorer.",
            "Permission Error","EyePort was unable to analyze the files because there was no permission to access them. Close any applications that may be using the files or try elevating EyePort as Administrator.","Failed to Analyze","EyePort was unable to analyze the files. The files may be corrupted containing data in an incorrect format. There could also be a problem with the algorithm. Please try again with different files and if the issue persists, contact the developer.","No Radar Violations detected",f"The participant has no radar violations. They have looked at the screen regularly within the set interval.",
            "Radar Violations detected","The participant has the following radar violations: ","No Export Folder","You have not provided an export folder to export to.\nEyePort will save the files in\n","EyePort was unable to export the files because there was no permission to write them. Close any applications that may be using the files or try elevating EyePort as Administrator.","Unknown Error","EyePort was unable to export the files. This may be due to a problem with the code. Please try again with different files and if the issue persists, contact the developer.",
            "Problem Opening User Manual","EyePort was not able to find the User Manual. The file may have been moved or deleted.","Missing Resources","EyePort cannot find some critical resource files in the install location. This may be due to a corruption or pirated download. Please re-install the EyePort software again.","Reset Settings","Are you sure you want to reset all settings?\nEyePort will apply the Default Settings.",
            "DynaFRAM not installed","EyePort has detected that DynaFRAM is not installed on your computer. Would you like EyePort to open the website for you to install it?","Glasses still connected","The TOBII Glasses are still connected to EyePort. It is recommended that you disconnect them before closing the program. Are you sure you still want to close EyePort?","Close Program Confirmation","Are you sure you want to close EyePort?","Memory Overload","There are too many Areas of Interests detected. The nature of rendering over 400 images at once will cause certain instability in EyePort. Please increase the Fixation Time in settings or record shorter sessions. It is a good idea to force quit EyePort at this time.",
            "Head Orientation Calibration","Choose the starting position of the participant's head orientation.","Head Up","Head Level","Head Down","Additional Resource Missing","Lite Version Installed. Please install EyePort Object Detection Module to use this feature.","Head Orientation Sensitivity","Choose how fast the participant's head movement should move for head orientation detections.","Radar Violation Interval","Choose the timing interval that defines how often the participant should look at the Radar Screen.","Saved Names Found","You have defined Object Names on this Project before. Use them now?",
            "Automatic Object Detection","Updating Interface","Progress","General","Basic","Advanced"] #167

    French = ["EyePort", "Version ", "Créé par Akash Samanta", "Fermer EyePort", "Exporter les Données", "Analyser les Données", "Lecture des Données", "Enregistrer les Données", "Paramètres", "Analyser les Données" ,"Enregistrer des Données","Exporter des Données","Bienvenue sur EyePort","Pour commencer, accédez à la page d'enregistrement des données.",
            "Manuel de l'utilisateur", "Ne plus afficher cette page au démarrage", "Utiliser les superpositions", "Taille de la boîte d'analyse", "Langue", "Thème", "Clair", "Sombre", "Système", "Mise à l'échelle de l'interface utilisateur" , "Retard de commutation homme mort", "Mode plein écran","Aperçu pop en direct","Réinitialiser tous les paramètres","Définir les zones interdites" ,"Définissez la longueur du côté du carré de tolérance en pixels. Cette image recadrée sera utilisée pour détecter les zones d'intérêt uniques.","Temps de fixation",
            "Choisissez combien de temps vous devez vous concentrer sur un objet avant de l'enregistrer comme zone d'intérêt.", "Sensibilité de la détection", "Choisissez la sensibilité de l'algorithme d'analyse pour détecter les zones d'intérêt uniques.","Détection automatique d'objet","Choisissez si vous souhaitez utiliser la détection automatique d'objets.","Désactivé","Général","Navires et Icebergs","VISTA Diesel Engine",
            "Mémoriser le dernier projet", "Compact Mode", "Choisissez le délai nécessaire pour activer le Dead Man Switch après aucune réponse du participant.", "Écraser les anciennes exportations", "Demander les noms de fichiers avant d'exporter", "Mode manuel", "Sélectionner le dossier de sortie", "Choisir où vous souhaitez exporter vos fichiers analysés.","Collecte de données sur le verre","Vérification des données de superposition du regard","Effacer la page","Commutateur d'homme mort",
            "Détections","Afficher la page de bienvenue","Chargement","Exporter le dossier","Exporter CSV","Exporter XMFV","Exporter XLSX","Afficher dans l'Explorateur","Ouvrir DynaFRAM","Charger le dossier de la carte SD","Chargez le dossier de projet souhaité à partir de la carte SD.","Démarrer l'analyse","Cliquez sur \"Analyser maintenant\" pour lancer l'analyse.",
            "Sélectionner un dossier", "Analyser Maintenant", "Temps d'orientation de la tête", "Temps tête haute", "Temps au niveau de la tête", "Temps tête basse", "Modifier les noms", "Violations de zone interdite", "Commutateur Homme Mort : Non Actif","Afficher les Graphiques","Violations Radar","Calcul des zones d'intérêt","Calcul des données d'orientation de la tête","Connecter les Lunettes","Synchronisation des horodatages" ,"Obtenir des données de trame","Calibrer les Lunettes","Calcul de zones d'intérêt uniques","Calcul des violations de zone interdite",
            "Enregistrer Maintenant", "Vue en direct des lunettes TOBII", "Non Disponible", "Déconnecter les Lunettes", "Démarrer la vue en direct", "Sélection Automatique", "Réussite de la Connexion", "Arrêter la vue en direct", "Chargement", "Arrêter Enregistrement","Erreur lors du démarrage de l'enregistrement","EyePort ne parvient pas à communiquer avec les lunettes TOBII. Peut-être vérifier à nouveau la connexion?","Erreur à l'arrêt de l'enregistrement","EyePort ne parvient pas à arrêter l'enregistrement. Cela pourrait être dû à un mauvais connexion avec les lunettes. Votre enregistrement continuera jusqu'à ce que les lunettes soient à court de batterie ou d'espace sur la carte SD. Appuyez sur le bouton d'alimentation pendant 4 secondes pour un arrêt contrôlé.",
            "Démarrer l'étalonnage", "Veuillez demander à l'utilisateur de regarder la carte d'étalonnage, puis ignorez cette invite pour démarrer l'étalonnage.", "Échec de l'étalonnage", "Échec de l'étalonnage", "Veuillez demander à l'utilisateur de continuer à regarder la carte pendant le processus d'étalonnage. Il n'y avait pas suffisamment d'échantillons valides pour terminer l'étalonnage. Veuillez réessayer.","Calibrer à Nouveau","Dossier de projet chargé","Utilisation de ","Veuillez charger le dossier de projet pour démarrer l'analyse.","Enregistrer Modifications","Zone d'intérêt unique : ","Zone d'intérêt : ","Commutateur Homme Mort : Actif",
            "Sélectionné", "Échec de la connexion", "EyePort ne parvient pas à se connecter aux lunettes TOBII.\nVeuillez vérifier la connexion et réessayer.", "Enregistrement toujours en cours", "Vous avez un enregistrement en cours. Veuillez arrêter l'enregistrement. puis déconnectez les lunettes.","Problème de chargement des fichiers","EyePort n'a pas pu localiser les fichiers requis dans le répertoire donné. Veuillez réessayer.","Échec de l'attribution de noms uniques","Vous ne pouvez pas attribuer des noms identiques aux zones uniques. d'intérêt. Utilisez le mode manuel pour ignorer ce comportement.","Problème de chargement des fichiers","EyePort n'a pas pu ouvrir les fichiers. Les fichiers ont peut-être été déplacés ou supprimés. Veuillez vérifier les fichiers chargés dans l'explorateur de fichiers.",
            "Erreur d'autorisation", "EyePort n'a pas pu analyser les fichiers car il n'y avait aucune autorisation pour y accéder. Fermez toutes les applications susceptibles d'utiliser les fichiers ou essayez d'élever EyePort en tant qu'administrateur.","Échec de l'analyse", "EyePort n'a pas pu pour analyser les fichiers. Les fichiers peuvent être corrompus et contenir des données dans un format incorrect. Il pourrait également y avoir un problème avec l'algorithme. Veuillez réessayer avec des fichiers différents et si le problème persiste, contactez le développeur.","Aucune violation radar détectée","Le participant n'a aucune violation du radar. Il a regardé l'écran régulièrement dans l'intervalle défini.",
            "Violations radar détectées", "Le participant présente les violations radar suivantes :", "Aucun dossier d'exportation", "Vous n'avez pas fourni de dossier d'exportation vers lequel exporter.\nEyePort enregistrera les fichiers dans\n", "EyePort n'a pas pu exporter les fichiers car il n'y avait pas d'autorisation pour les écrire. Fermez toutes les applications susceptibles d'utiliser les fichiers ou essayez d'élever EyePort en tant qu'administrateur.","Erreur inconnue","EyePort n'a pas pu exporter les fichiers. Ceci peut être dû à un problème avec le code. Veuillez réessayer avec des fichiers différents et si le problème persiste, contactez le développeur.",
            "Problème d'ouverture du manuel de l'utilisateur", "EyePort n'a pas pu trouver le manuel de l'utilisateur. Le fichier a peut-être été déplacé ou supprimé.", "Ressources manquantes", "EyePort ne peut pas trouver certains fichiers de ressources critiques dans l'emplacement d'installation. Cela peut être en raison d'un téléchargement corrompu ou piraté. Veuillez réinstaller le logiciel EyePort.","Réinitialiser les paramètres","Êtes-vous sûr de vouloir réinitialiser tous les paramètres?\nEyePort appliquera les paramètres par défaut.",
            "DynaFRAM non Installé", "EyePort a détecté que DynaFRAM n'est pas installé sur votre ordinateur. Souhaitez-vous qu'EyePort ouvre le site Web pour que vous puissiez l'installer?", "Les lunettes sont toujours connectées", "Les lunettes TOBII sont toujours connectées à EyePort . Il est recommandé de les déconnecter avant de fermer le programme. Êtes-vous sûr de vouloir toujours quitter EyePort?","Fermer la confirmation du programme", "Êtes-vous sûr de vouloir fermer EyePort?","Surcharge de mémoire", "Il y a trop de zones d'intérêt détectées. La nature du rendu de plus de 400 images à la fois entraînera une certaine instabilité dans EyePort. Veuillez augmenter le temps de fixation dans les paramètres ou enregistrer des sessions plus courtes. C'est une bonne idée de forcer quittez EyePort à ce moment-là.",
            "Calibrage de l'orientation de la tête", "Choisissez la position de départ de l'orientation de la tête du participant.", "Tête Haute", "Tête Droit", "Tête Baissée","Ressource supplémentaire manquante", "Version allégée installée. Veuillez installer le module de détection d'objets EyePort pour utiliser cette fonctionnalité.","Sensibilité d'orientation de la tête", "Choisissez la vitesse à laquelle le mouvement de la tête du participant doit se déplacer pour les détections d’orientation de la tête.","Intervalle de violation du radar","Choisissez l'intervalle de temps qui définit la fréquence à laquelle le participant doit regarder l'écran radar.","Noms enregistrés trouvés", "Vous avez déjà défini des noms d'objet sur ce projet. Utilisez-les maintenant?",
            "Détection automatique d'objets", "Mise à jour de l'interface","Progrès","Général", "Basique", "Avancé"]

    Dutch = ["EyePort", "Versie ", "Gemaakt door Akash Samanta", "Sluit EyePort", "Gegevens Exporteren", "Data Analyseren", "Gegevens Afspelen", "Gegevens Opnemen", "Instellingen", "Data Analyseren","Gegevens Opnemen","Gegevens Exporteren","Welkom bij EyePort","Om aan de slag te gaan, gaat u naar de pagina \"Gegevens Opnemen\".",
            "Handleiding", "Deze pagina niet meer tonen bij opstarten", "Overlays gebruiken", "Grootte analysebox", "Taal", "Thema", "Licht", "Donker", "Systeem", "UI-Schaling" ,"Dodemansknop vertraging","Volledig Schermmodus", "Pop Live View", "Reset alle instellingen", "Definieer No Go-zones", "Definieer de tolerantie van de vierkante zijdelengte in pixels. Deze bijgesneden afbeelding wordt gebruikt voor het detecteren van unieke interessegebieden.", "Fixatietijd",
            "Kies hoe lang u zich op een object wilt concentreren voordat u het als interessegebied registreert.","Detectiegevoeligheid", "Kies hoe gevoelig het analysealgoritme moet zijn om unieke interessegebieden te detecteren.","Automatische objectdetectie", "Kies of u automatische objectdetectie wilt gebruiken.", "Uitgeschakeld", "Algemeen", "Schepen en ijsbergen", "VISTA-dieselmotor",
            "Onthoud het laatste project", "Compacte Modus", "Kies de tijdsvertraging die nodig is om de dodemansknop te activeren nadat de deelnemer geen reactie heeft gegeven.", "Oude exports overschrijven", "Vraag om bestandsnamen voordat u gaat exporteren", "Handmatige modus", "Selecteer de uitvoermap", "Kies waar u wilt uw geanalyseerde bestanden exporteren.","Glasgegevens verzamelen","Gaze Overlay-gegevens controleren","Pagina wissen","Dodemansschakelaar",
            "Detecties","Welkomstpagina weergeven", "Analyseren","Map Exporteren", "CSV Exporteren", "XMFV Exporteren", "XLSX Exporteren", "Weergeven in Verkenner", "DynaFRAM openen", "SD-kaartmap laden", "Laad de gewenste projectmap vanaf de SD-kaart.","Start Analyse","Klik op \"Analyseer Nu\" om te beginnen met analyseren.",
            "Selecteer Folder", "Analyseer Nu", "Hoofdoriëntatietijden", "Head-up-tijd", "Hoofdniveau-tijd", "Head-down-tijd", "Namen bewerken", "No-Go Zone-overtredingen", "Dode man Schakelaar: Niet Actief","Grafieken weergeven","Radarovertredingen","Belangengebieden berekenen","Hoofdoriëntatiegegevens berekenen","Bril Verbinden", "Tijdstempels synchroniseren","Framegegevens ophalen","Bril Kalibreren", "Unieke interessegebieden berekenen", "Berekening van overtredingen in de No-Go-zone",
            'Nu Opnemen', 'TOBII Bril Live View', 'Niet beschikbaar', 'Bril loskoppelen', 'Liveweergave starten', 'Automatisch geselecteerd', 'Verbinding geslaagd', 'Liveweergave stoppen', 'Laden', "Stoppen Opname" ,"Fout bij het starten van de opname","EyePort kan niet communiceren met de TOBII-bril. Controleer de verbinding misschien nog eens?","Fout bij het stoppen van de opname","EyePort kan de opname niet stoppen. Dit kan te wijten zijn aan een slechte verbinding met de bril. Uw opname gaat door totdat de batterij van de bril of de SD-kaart leeg is. Houd de aan/uit-knop 4 seconden ingedrukt voor een gecontroleerde uitschakeling.",
            "Start Kalibratie", "Vraag de gebruiker om naar de kalibratiekaart te kijken en negeer vervolgens deze prompt om de kalibratie te starten.","Kalibratie Mislukt", "Kalibratie Mislukt", "Vraag de gebruiker om tijdens de kalibratie naar de kaart te blijven kijken kalibratieproces. Er waren niet genoeg geldige monsters om de kalibratie te voltooien. Probeer het opnieuw.","Opnieuw kalibreren", "Projectmap geladen", "Gebruikt "," Laad de projectmap om de analyse te starten.", "Opslaan Veranderingen ","Uniek interessegebied: ","Interessegebied: ","Dead Man Switch: Actief",
            "Geselecteerd", "Verbinding Mislukt", "EyePort kan geen verbinding maken met de TOBII-bril.\nControleer de verbinding en probeer het opnieuw.", "Opname nog steeds bezig", "U bent bezig met opnemen. Stop de opname. en koppel vervolgens de bril los.","Probleem met het laden van bestanden", "EyePort kon de vereiste bestanden niet vinden in de opgegeven map. Probeer het opnieuw.","Kan geen unieke namen toewijzen", "U kunt geen identieke namen toewijzen aan unieke gebieden van belang. Gebruik de handmatige modus om dit gedrag te onderdrukken.","Probleem bij het laden van bestanden", "EyePort kon de bestanden niet openen. De bestanden zijn mogelijk verplaatst of verwijderd. Controleer de geladen bestanden in de bestandsverkenner.",
            "Toestemmingsfout", "EyePort kon de bestanden niet analyseren omdat er geen toestemming was om er toegang toe te krijgen. Sluit alle toepassingen die mogelijk gebruik maken van de bestanden of probeer EyePort als beheerder te verhogen.","Kan niet analyseren", "EyePort kon de bestanden niet analyseren. De bestanden zijn mogelijk beschadigd en bevatten gegevens in een onjuist formaat. Er kan ook een probleem zijn met het algoritme. Probeer het opnieuw met andere bestanden en neem contact op met de ontwikkelaar als het probleem zich blijft voordoen.","Geen radarovertredingen gedetecteerd","De deelnemer heeft geen radarovertredingen. Hij heeft regelmatig naar het scherm gekeken binnen de ingestelde interval.",
            "Radarschendingen gedetecteerd", "De deelnemer heeft de volgende radarschendingen: ","Geen exportmap", "U heeft geen exportmap opgegeven om naar te exporteren.\nEyePort zal de bestanden opslaan in\n", "EyePort was niet in staat om de bestanden te exporteren, omdat er geen toestemming was om ze te schrijven. Sluit alle toepassingen die mogelijk gebruik maken van de bestanden of probeer EyePort als beheerder te verhogen.","Onbekende fout","EyePort kon de bestanden niet exporteren. Dit kan te wijten zijn aan een probleem met de code. Probeer het opnieuw met andere bestanden en neem contact op met de ontwikkelaar als het probleem zich blijft voordoen.",
            "Probleem bij het openen van de Handleiding", "EyePort kon de Handleiding niet vinden. Het bestand is mogelijk verplaatst of verwijderd.","Ontbrekende bronnen", "EyePort kan sommige kritieke bronbestanden niet vinden op de installatielocatie. Dit kan het geval zijn vanwege een corruptie of een illegale download. Installeer de EyePort-software opnieuw.","Reset instellingen","Weet u zeker dat u alle instellingen wilt resetten?\nEyePort zal de standaardinstellingen toepassen.",
            "DynaFRAM niet geïnstalleerd", "EyePort heeft gedetecteerd dat DynaFRAM niet op uw computer is geïnstalleerd. Wilt u dat EyePort de website opent zodat u deze kunt installeren?", "Glazen nog steeds verbonden", "De TOBII-bril is nog steeds verbonden met EyePort . Het wordt aanbevolen dat u de verbinding verbreekt voordat u het programma afsluit. Weet u zeker dat u EyePort nog steeds wilt afsluiten? ", "Programmabevestiging sluiten", "Weet u zeker dat u EyePort wilt sluiten?","Geheugenoverbelasting", "Er zijn te veel interessegebieden gedetecteerd. De aard van het tegelijk weergeven van meer dan 400 afbeeldingen zal een zekere instabiliteit in EyePort veroorzaken. Verhoog de fixatietijd in de instellingen of neem kortere sessies op. Het is een goed idee om te forceren sluit EyePort nu af.",
            "Hoofdoriëntatie kalibratie", "Kies de startpositie voor de hoofdoriëntatie van de deelnemer.", "Hoofd Omhoog  ", "Hoofd Waterpas", "Hoofd Omlaag   ","Extra bronnen ontbreken", "Lite-versie geïnstalleerd. Installeer de EyePort Object Detection Module om deze functie te gebruiken.","Hoofdoriëntatiegevoeligheid", "Kies hoe snel de hoofdbeweging van de deelnemer moet bewegen voor detectie van hoofdoriëntatie.","Radarschendingsinterval","Kies het tijdsinterval dat de frequentie van de deelnemer met betrekking tot de radar bepaalt.","Opgeslagen namen gevonden", "U heeft al eerder objectnamen voor dit project gedefinieerd. Gebruik ze nu?",
            "Automatische objectdetectie", "Interface bijwerken","Voortgang","Algemeen", "Basis", "Geavanceerd"]

    if language_mode==0:
        lang = English
    elif language_mode==1:
        lang = French
    else:
        lang = Dutch

    #Language Check
    # for i in range(len(English)):
    #     try:
    #         print(i,English[i],"\n",i,Dutch[i],sep="",end="\n")
    #     except:
    #         pass
    # quit()

    #Frames
    frame1=CTkFrame(root,fg_color=("#cfcfcf","#333333")) #Left Bar
    analyzeframe=CTkScrollableFrame(root)
    viewframe=CTkFrame(root)
    exportframe=CTkFrame(root)
    recordframe=CTkFrame(root)
    welcomeframe=CTkFrame(root)
    settingsframe=CTkScrollableFrame(root) 
    videoframe = CTkFrame(recordframe)
    progressframe = CTkFrame(analyzeframe,fg_color=("#cfcfcf","#333333"))
    imgframe = CTkScrollableFrame(analyzeframe,fg_color=("#cfcfcf","#333333"),orientation="horizontal",height=270)
    imgframe2 = CTkScrollableFrame(analyzeframe,fg_color=("#cfcfcf","#333333"),orientation="horizontal",height=270)
    imgframe3 = CTkFrame(analyzeframe)
    imgframe6 = CTkFrame(analyzeframe)
    imgframe4 = CTkFrame(analyzeframe,fg_color=("#cfcfcf","#333333"))
    imgframe5 = CTkScrollableFrame(analyzeframe,fg_color=("#cfcfcf","#333333"),orientation="horizontal",height=40)
    imgframe4.columnconfigure((0,1,2),weight=1)
    expframe = CTkFrame(exportframe)
    expframe1 = CTkFrame(exportframe)
    expframe1.columnconfigure((0,1,2),weight=1)

    #Button Scaling
    bsize=16
    bheight=35
    bwidth=200
    bsheight=40
    bswidth=40

    #Images
    try:
        image1 = Image.open(".\Resources\Manual.png")
        image2 = Image.open(".\Resources\Record.png")
        image3 = Image.open(".\Resources\Analyze.png")
        image4 = Image.open(".\Resources\Export.png")
        image5 = Image.open(".\Resources\Settings.png")
        image6 = Image.open(".\Resources\HeadUp.png")
        image7 = Image.open(".\Resources\Manual1.png")
        image8 = Image.open(".\Resources\Manual2.png")
        image11 = Image.open(".\Resources\FullBat.png")
        image12 = Image.open(".\Resources\QFBat.png")
        image13 = Image.open(".\Resources\HalfBat.png")
        image14 = Image.open(".\Resources\LowBat.png")
        image15 = Image.open(".\Resources\ZeroBat.png")
        image16 = Image.open(".\Resources\Timeleft.png")
        image17 = Image.open(".\Resources\EyeLight.png")
        image18 = Image.open(".\Resources\EyeDark.png")
        image19 = Image.open(".\Resources\Exit.png")
        image20 = Image.open(".\Resources\Pause.png")
        image21 = Image.open(".\Resources\ExportSuccess.png")
        image22 = Image.open(".\Resources\DExportSuccess.png")
        image23 = Image.open(".\Resources\ExportFailed.png")
        image24 = Image.open(".\Resources\DExportFailed.png")
        image25 = Image.open(".\Resources\ExportClick.png")
        image26 = Image.open(".\Resources\DExportClick.png")
        image27 = Image.open(".\Resources\ConnectGlasses.png")
        image28 = Image.open(".\Resources\PreviewPop.png")
        image29 = Image.open(".\Resources\PreviewAvail.png")
        image291 = Image.open(".\Resources\PreviewLoading.png")
        image30 = Image.open(".\Resources\HeadLevel.png")
        image31 = Image.open(".\Resources\HeadDown.png")
        image32 = Image.open(".\Resources\HeadUpDark.png")
        image33 = Image.open(".\Resources\HeadDownDark.png")
        image34 = Image.open(".\Resources\HeadLevelDark.png")
        image35 = Image.open(".\Resources\Play.png")
        image36 = Image.open(".\Resources\Back.png")
        image37 = Image.open(".\Resources\Forward.png")
        image38 = Image.open(".\Resources\RecordProgress.png")
        image39 = Image.open(".\Resources\DRecordProgress.png")
        image40 = Image.open(".\Resources\RecordSaved.png")
        image41 = Image.open(".\Resources\DRecordSaved.png")
        image42 = Image.open(".\Resources\Logo2.png")
        image43 = Image.open(".\Resources\combo1.png")
        image44 = Image.open(".\Resources\combo2.png")
        image45 = Image.open(".\Resources\combo3.png")
        image46 = Image.open(".\Resources\combo4.png")
        image47 = Image.open(".\Resources\combo41.png")
        image48 = Image.open(".\Resources\combo5.png")
        image49 = Image.open(".\Resources\combo6.png")
        image50 = Image.open(".\Resources\combo71.png")
        image51 = Image.open(".\Resources\combo72.png")
        image52 = Image.open(".\Resources\combo8.png")
        image53 = Image.open(".\Resources\combo81.png")
        image54 = Image.open(".\Resources\combo9.png")
        image55 = Image.open(".\Resources\combo91.png")
        image56 = Image.open(".\Resources\combo10.png")
        image57 = Image.open(".\Resources\combo101.png")
        image58 = Image.open(".\Resources\combo51.png")
        image59 = Image.open(".\Resources\combo42.png")
        image60 = Image.open(".\Resources\combo43.png")
        image61 = Image.open(".\Resources\combo73.png")
        image62 = Image.open(".\Resources\combo74.png")
        image63 = Image.open(".\Resources\combo61.png")
        image64 = Image.open(".\Resources\Tolerance\Tolerance10.png")
        image65 = Image.open(".\Resources\Tolerance\Tolerance11.png")
        image66 = Image.open(".\Resources\Tolerance\Tolerance20.png")
        image67 = Image.open(".\Resources\Tolerance\Tolerance21.png")
        image68 = Image.open(".\Resources\Tolerance\Tolerance30.png")
        image69 = Image.open(".\Resources\Tolerance\Tolerance31.png")
        image70 = Image.open(".\Resources\Tolerance\Tolerance40.png")
        image71 = Image.open(".\Resources\Tolerance\Tolerance41.png")
        image72 = Image.open(".\Resources\Tolerance\Tolerance50.png")
        image73 = Image.open(".\Resources\Tolerance\Tolerance51.png")
        image74 = Image.open(".\Resources\Tolerance\Tolerance60.png")
        image75 = Image.open(".\Resources\Tolerance\Tolerance61.png")
        image76 = Image.open(".\Resources\Tolerance\Tolerance70.png")
        image77 = Image.open(".\Resources\Tolerance\Tolerance71.png")
        image78 = Image.open(".\Resources\Tolerance\Tolerance80.png")
        image79 = Image.open(".\Resources\Tolerance\Tolerance81.png")
        image80 = Image.open(".\Resources\Tolerance\Tolerance90.png")
        image81 = Image.open(".\Resources\Tolerance\Tolerance91.png")
        image82 = Image.open(".\Resources\Tolerance\Tolerance100.png")
        image83 = Image.open(".\Resources\Tolerance\Tolerance101.png")
        image84 = Image.open(".\Resources\Tolerance\Tolerance110.png")
        image85 = Image.open(".\Resources\Tolerance\Tolerance111.png")
        image86 = Image.open(".\Resources\Tolerance\Tolerance120.png")
        image87 = Image.open(".\Resources\Tolerance\Tolerance121.png")
        image88 = Image.open(".\Resources\Tolerance\Tolerance130.png")
        image89 = Image.open(".\Resources\Tolerance\Tolerance131.png")
    except:
        tk.messagebox.showerror(lang[137],lang[138]) #Missing Resources
        quit()
    imgtk1 = CTkImage(light_image=image1)
    imgtk2 = CTkImage(light_image=image2)
    imgtk3 = CTkImage(light_image=image3)
    imgtk4 = CTkImage(light_image=image4)
    imgtk5 = CTkImage(light_image=image5)
    imgtk6 = CTkImage(light_image=image6,dark_image=image32,size=(135,146))
    imgtk9 = CTkImage(light_image=image11)
    imgtk10 = CTkImage(light_image=image12)
    imgtk11 = CTkImage(light_image=image13)
    imgtk12 = CTkImage(light_image=image14)
    imgtk13 = CTkImage(light_image=image15)
    imgtk14 = CTkImage(light_image=image16)
    imgtk15 = CTkImage(light_image=image17,dark_image=image18,size=(400,400))
    imgtk151 = CTkImage(light_image=image35)
    imgtk16 = CTkImage(light_image=image19)
    imgtk18 = CTkImage(light_image=image21,dark_image=image22,size=(270,200))
    imgtk19 = CTkImage(light_image=image23,dark_image=image24,size=(228,200))
    imgtk20 = CTkImage(light_image=image25,dark_image=image26,size=(209,200))
    imgtk21 = CTkImage(light_image=image43,size=(180,110))
    imgtk22 = CTkImage(light_image=image44,size=(180,110))
    imgtk23 = CTkImage(light_image=image45,size=(180,110))
    imgtk24 = CTkImage(light_image=image58,size=(143,88))
    imgtk25 = CTkImage(light_image=image59,dark_image=image60,size=(143,88))
    imgtk26 = CTkImage(light_image=image38,dark_image=image39,size=(170,30))
    imgtk27 = CTkImage(light_image=image40,dark_image=image41,size=(150,30))
    imgtk28 = CTkImage(light_image=image20)
    imgtk29 = CTkImage(light_image=image61,dark_image=image62,size=(143,88))
    imgtk30 = CTkImage(light_image=image30,dark_image=image34,size=(115,146))
    imgtk31 = CTkImage(light_image=image31,dark_image=image33,size=(140,146))
    imgtk32 = CTkImage(light_image=image52,dark_image=image53,size=(180,110))
    imgtk33 = CTkImage(light_image=image54,dark_image=image55,size=(180,110))
    imgtk34 = CTkImage(light_image=image56,dark_image=image57,size=(180,110))
    imgtk35 = CTkImage(light_image=image63,size=(180,110))
    imgtk36 = CTkImage(light_image=image36)
    imgtk37 = CTkImage(light_image=image37)
    imgtk38 = CTkImage(light_image=image46,dark_image=image47,size=(180,110))
    imgtk39 = CTkImage(light_image=image48,size=(180,110))
    imgtk40 = CTkImage(light_image=image49,size=(143,88))
    imgtk41 = CTkImage(light_image=image50,dark_image=image51,size=(180,110))
    imgtk42 = CTkImage(light_image=image42,size=(60,60))
    imgtk43 = CTkImage(light_image=image64,dark_image=image65,size=(236,145))
    imgtk44 = CTkImage(light_image=image66,dark_image=image67,size=(236,145))
    imgtk45 = CTkImage(light_image=image68,dark_image=image69,size=(236,145))
    imgtk46 = CTkImage(light_image=image70,dark_image=image71,size=(236,145))
    imgtk47 = CTkImage(light_image=image72,dark_image=image73,size=(236,145))
    imgtk48 = CTkImage(light_image=image74,dark_image=image75,size=(236,145))
    imgtk49 = CTkImage(light_image=image76,dark_image=image77,size=(236,145))
    imgtk50 = CTkImage(light_image=image78,dark_image=image79,size=(236,145))
    imgtk51 = CTkImage(light_image=image80,dark_image=image81,size=(236,145))
    imgtk52 = CTkImage(light_image=image82,dark_image=image83,size=(236,145))
    imgtk53 = CTkImage(light_image=image84,dark_image=image85,size=(236,145))
    imgtk54 = CTkImage(light_image=image86,dark_image=image87,size=(236,145))
    imgtk55 = CTkImage(light_image=image88,dark_image=image89,size=(236,145))
    imgtk59 = CTkImage(light_image=image7,dark_image=image8)

    #Left Bar Elements
    text80 = CTkButton(frame1, text="", image = imgtk42,command=lambda :webbrowser.open("https://www.akashcraft.ca/eyeport.html"),hover=False, fg_color="transparent")
    heading1 = CTkLabel(frame1, text=lang[0], font=CTkFont(size=30)) #EyePort
    if imageai_avail:
        heading2 = CTkLabel(frame1, text=lang[1]+"3.3.3 (Full)", font=CTkFont(size=15)) #Version
    else:
        heading2 = CTkLabel(frame1, text=lang[1]+"3.3.3 (Lite)", font=CTkFont(size=15)) #Version
    text1 = CTkLabel(frame1, text=lang[2]) #Made by Akash Samanta
    quitbutton = CTkButton(frame1,text=lang[3],command=destroyer, font=CTkFont(size=bsize),width=bwidth,height=bheight,fg_color="dark red", image=imgtk16,hover_color='#530000')
    exportdata = CTkButton(frame1,text=lang[4],command=exportpage, font=CTkFont(size=bsize),width=bwidth,height=bheight, image=imgtk4)
    analyzedata = CTkButton(frame1,text=lang[5],command=analyzepage, font=CTkFont(size=bsize),width=bwidth,height=bheight, image=imgtk3)
    viewdata = CTkButton(frame1,text=lang[6],command=viewpage, font=CTkFont(size=bsize),width=bwidth,height=bheight, image=imgtk151,state='disabled')
    recorddata = CTkButton(frame1,text=lang[7],command=recordpage, font=CTkFont(size=bsize),width=bwidth,height=bheight, image=imgtk2)
    settingsbutton = CTkButton(frame1,text=lang[8],command=settingspage, font=CTkFont(size=bsize),width=bwidth,height=bheight, image=imgtk5)
    usermanual2 = CTkButton(frame1,text=lang[14],command=showusermanual, font=CTkFont(size=bsize),width=bwidth,height=bheight, image=imgtk1)

    text2 = CTkLabel(analyzeframe, text=lang[9], font=CTkFont(size=25))
    text3 = CTkLabel(recordframe, text=lang[10], font=CTkFont(size=25))
    text4 = CTkLabel(exportframe, text=lang[11], font=CTkFont(size=25))

    #Welcome Elements
    text37 = CTkLabel(welcomeframe, text="", image = imgtk15)
    text38 = CTkLabel(welcomeframe, text=lang[12], font=CTkFont(size=45)) #Welcome to EyePort
    text39 = CTkLabel(welcomeframe, text=lang[13], font=CTkFont(size=15)) #Desc
    gorecbutton = CTkButton(welcomeframe,text=lang[7],command=recordpage, font=CTkFont(size=bsize),width=bwidth,height=bheight,image=imgtk2)
    usermanual = CTkButton(welcomeframe,text=lang[14],command=showusermanual, font=CTkFont(size=bsize),width=bwidth,height=bheight, image=imgtk1)
    showagain_var = tk.IntVar(value=0)
    showagain = CTkCheckBox(welcomeframe, text=lang[15], command=showagain,variable=showagain_var, onvalue=1, offvalue=0) #Never Show Again

    #View Elements
    overlay_var = tk.IntVar(value=overlay)
    check_1 = CTkSwitch(viewframe, text=lang[16], variable=overlay_var, onvalue=1, offvalue=0, font=CTkFont(size=15),command=toggleoverlay) #Use Overlays
    text36 =  CTkLabel(viewframe, text=lang[6], font=CTkFont(size=25)) #Playback Data
    viewframe1 = CTkFrame(viewframe)
    playpausebutton = CTkButton(viewframe1,text="",width=25,height=30,image=imgtk151,command=playvideo,state='disabled')
    forwardbutton = CTkButton(viewframe1,text="",width=20,height=25,image=imgtk36,state='disabled',command=lambda: skip(-1))
    backbutton = CTkButton(viewframe1,text="",width=20,height=25,image=imgtk37,state='disabled',command=lambda: skip(1))
    viewstart = CTkLabel(viewframe1, text="0:00", font=CTkFont(size=12))
    viewend = CTkLabel(viewframe1, text="0:00", font=CTkFont(size=12))
    viewslider = CTkSlider(master=viewframe1, from_=0, to=100, state='disabled',command=seek)
    viewslider.set(0)
    player = TkinterVideo(scaled=True,master=viewframe1)
    player.bind("<<Duration>>", update_duration)
    player.bind("<<SecondChanged>>", update_scale)
    player.bind("<<Ended>>", video_ended)
    viewtext = CTkLabel(viewframe1, text=" Object 1 ",font=CTkFont(size=15))

    #Settings Elements
    settingstab = CTkTabview(settingsframe,fg_color=("#cfcfcf","#333333"))
    settingstab.add(lang[165])
    settingstab.add(lang[166])
    settingstab.add(lang[167])
    frame2 = CTkFrame(settingstab.tab(lang[165])) #Theme
    frame3 = CTkFrame(settingstab.tab(lang[165])) #Language
    frame4 = CTkFrame(settingstab.tab(lang[166])) #Head Orientation
    frame6 = CTkFrame(settingstab.tab(lang[166])) #Object Detection
    text5 = CTkLabel(settingsframe, text=lang[8], font=CTkFont(size=25)) #Settings
    text71 = CTkLabel(settingstab.tab(lang[165]), text=lang[18], font=CTkFont(size=15)) #Language
    lang_var = tk.IntVar(value=value[24])
    radio_button_4 = CTkRadioButton(frame3, text="English", variable=lang_var, value=0, command=updatelang)
    radio_button_5 = CTkRadioButton(frame3, text="Français", variable=lang_var, value=1, command=updatelang)
    radio_button_6 = CTkRadioButton(frame3, text="Dutch", variable=lang_var, value=2, command=updatelang)
    text7 = CTkLabel(settingstab.tab(lang[165]), text=lang[19], font=CTkFont(size=15)) #Theme
    if value[2][:-1]=="Light":
        theme = 0
    elif value[2][:-1]=="Dark":
        theme = 1
    else:
        theme = 2
    theme_var = tk.IntVar(value=theme)
    combolist = [imgtk21,imgtk22,imgtk23,imgtk38,imgtk39,imgtk35,imgtk41,imgtk32,imgtk33,imgtk34]
    if int(value[3][:-1])>=100:
        for i in combolist:
            i.configure(size=(144,88))
    combo1 = CTkLabel(frame2,text='',image=imgtk21)
    combo2 = CTkLabel(frame2,text='',image=imgtk22)
    combo3 = CTkLabel(frame2,text='',image=imgtk23)
    combo4 = CTkLabel(frame6,text='',image=imgtk38)
    combo5 = CTkLabel(frame6,text='',image=imgtk39)
    combo6 = CTkLabel(frame6,text='',image=imgtk35)
    combo7 = CTkLabel(frame6,text='',image=imgtk41)
    combo8 = CTkLabel(frame4,text='',image=imgtk32)
    combo9 = CTkLabel(frame4,text='',image=imgtk33)
    combo10 = CTkLabel(frame4,text='',image=imgtk34)
    combo11 = CTkLabel(settingstab.tab(lang[167]),text='',image=imgtk43)
    combo12 = CTkLabel(settingstab.tab(lang[167]),text='',image=imgtk55)
    radio_button_1 = CTkRadioButton(frame2, text=lang[20], variable=theme_var, value=0, command=updatetheme) #Light
    radio_button_2 = CTkRadioButton(frame2, text=lang[21], variable=theme_var, value=1, command=updatetheme) #Dark
    radio_button_3 = CTkRadioButton(frame2, text=lang[22], variable=theme_var, value=2, command=updatetheme) #System
    text8 = CTkLabel(settingstab.tab(lang[165]), text=lang[23], font=CTkFont(size=15)) #UI Scaling
    scaling_optionmenu = CTkOptionMenu(settingstab.tab(lang[165]), values=["80%", "90%", "100%", "110%", "120%"],command=updatescale)
    scaling_optionmenu.set(value[3][:-1]+"%")
    fullscr_var = tk.IntVar(value=fullscreenmode)
    check_3 = CTkCheckBox(settingstab.tab(lang[165]), text=lang[25], variable=fullscr_var, onvalue=1, offvalue=0, font=CTkFont(size=15),command=usefullscreen) #Start in Fullscreen Mode
    reset = CTkButton(settingstab.tab(lang[165]),text=lang[27],font=CTkFont(size=bsize),width=bwidth,height=bheight,command=resetsettings2) #Reset All Settings
    nogozone = CTkButton(settingstab.tab(lang[166]),text=lang[28],font=CTkFont(size=bsize),width=bwidth,height=bheight,command=nogozonefun) #Define No Go Zones
    text68 = CTkLabel(settingstab.tab(lang[166]), text=lang[30], font=CTkFont(size=15)) #Fixation Time
    text68help = CTkButton(settingstab.tab(lang[166]),text="",image=imgtk59,width=10,height=10,fg_color="transparent",hover=False,command=lambda: tk.messagebox.showinfo(lang[30],lang[31]))
    slider_2 = CTkSlider(settingstab.tab(lang[166]), width=200, from_=25, to=500, number_of_steps=19, command=updatetimehit) 
    slider_2.set(int(value[17][:-1]))
    if thit%50==0:
        text70 = CTkLabel(settingstab.tab(lang[166]), text=str(thit/100)+'0 s', font=CTkFont(size=12))
    else:
        text70 = CTkLabel(settingstab.tab(lang[166]), text=str(thit/100)+' s', font=CTkFont(size=12))
    text81 = CTkLabel(settingstab.tab(lang[166]), text=lang[24], font=CTkFont(size=15)) #Dead Man Switch
    text81help = CTkButton(settingstab.tab(lang[166]),text="",image=imgtk59,width=10,height=10,fg_color="transparent",hover=False,command=lambda: tk.messagebox.showinfo(lang[24],lang[42]))
    slider_1 = CTkSlider(settingstab.tab(lang[166]), width=200, from_=10, to=120, number_of_steps=11, command=updateadead)
    slider_1.set(int(value[6][:-1]))
    text82 = CTkLabel(settingstab.tab(lang[166]), text=str(adead)+' s', font=CTkFont(size=12))
    text47 = CTkLabel(settingstab.tab(lang[167]), text=lang[32], font=CTkFont(size=15)) #Detection Sensitivity
    text47help = CTkButton(settingstab.tab(lang[167]),text="",image=imgtk59,width=10,height=10,fg_color="transparent",hover=False,command=lambda: tk.messagebox.showinfo(lang[32],lang[33]))
    scaling_optionmenu2 = CTkOptionMenu(settingstab.tab(lang[167]), values=["Extremely Low","Very Low","Low","Normal","High","Very High"], command=updateasense)
    scaling_optionmenu2.set(value[12][:-1])
    text72 = CTkLabel(settingstab.tab(lang[166]), text=lang[149], font=CTkFont(size=15)) #Head Orientation Calibration
    text72help = CTkButton(settingstab.tab(lang[166]),text="",image=imgtk59,width=10,height=10,fg_color="transparent",hover=False,command=lambda: tk.messagebox.showinfo(lang[149],lang[150]))
    headmode_var = tk.IntVar(value=int(value[20][:-1]))
    radio_button_12 = CTkRadioButton(frame4, text=lang[151], variable=headmode_var, value=1, command=updateheadmode)
    radio_button_13 = CTkRadioButton(frame4, text=lang[152], variable=headmode_var, value=2, command=updateheadmode)
    radio_button_14 = CTkRadioButton(frame4, text=lang[153], variable=headmode_var, value=3, command=updateheadmode)
    text74 = CTkLabel(settingstab.tab(lang[167]), text=lang[156], font=CTkFont(size=15)) #Head Acceleration
    text74help = CTkButton(settingstab.tab(lang[167]),text="",image=imgtk59,width=10,height=10,fg_color="transparent",hover=False,command=lambda: tk.messagebox.showinfo(lang[156],lang[157]))
    slider_3 = CTkSlider(settingstab.tab(lang[167]), width=200, from_=5, to=80, number_of_steps=15, command=updateheadsen)
    slider_3.set(int(value[4][:-1]))
    updateheadsenpic()
    text76 = CTkLabel(settingstab.tab(lang[167]), text=str(headsen)+' deg/s', font=CTkFont(size=12))
    text78 = CTkLabel(settingstab.tab(lang[167]), text=lang[17], font=CTkFont(size=15)) #Tolerance Box Size
    text78help = CTkButton(settingstab.tab(lang[167]),text="",image=imgtk59,width=10,height=10,fg_color="transparent",hover=False,command=lambda: tk.messagebox.showinfo(lang[17],lang[29]))
    slider_5 = CTkSlider(settingstab.tab(lang[167]), width=200, from_=50, to=1000, number_of_steps=19, command=updateatolbox)
    slider_5.set(int(value[8][:-1]))
    text83 = CTkLabel(settingstab.tab(lang[167]), text=str(atolbox)+' px', font=CTkFont(size=12))
    updatetolpic()
    text77 = CTkLabel(settingstab.tab(lang[166]), text=lang[158], font=CTkFont(size=15)) #Radar
    text77help = CTkButton(settingstab.tab(lang[166]),text="",image=imgtk59,width=10,height=10,fg_color="transparent",hover=False,command=lambda: tk.messagebox.showinfo(lang[158],lang[159]))
    slider_4 = CTkSlider(settingstab.tab(lang[166]), width=200, from_=5, to=80, number_of_steps=15, command=updateaviolate)
    slider_4.set(int(value[5][:-1]))
    text79 = CTkLabel(settingstab.tab(lang[166]), text=str(aviolate)+' s', font=CTkFont(size=12))
    useai_var = tk.IntVar(value=int(value[13][:-1]))
    text65 = CTkLabel(settingstab.tab(lang[166]), text=lang[34], font=CTkFont(size=15)) #Automatic Object Detection
    radio_button_8 = CTkRadioButton(frame6, text=lang[36], variable=useai_var, value=0, command=updateuseai)
    radio_button_9 = CTkRadioButton(frame6, text=lang[37], variable=useai_var, value=1, command=updateuseai)
    radio_button_10 = CTkRadioButton(frame6, text=lang[38], variable=useai_var, value=2, command=updateuseai)
    radio_button_11 = CTkRadioButton(frame6, text=lang[39], variable=useai_var, value=3, command=updateuseai)
    if not imageai_avail:
        text66 = CTkLabel(settingstab.tab(lang[166]), text=lang[155], font=CTkFont(size=12), text_color=("red","yellow")) #Desc
        radio_button_8.configure(state="disabled")
        radio_button_9.configure(state="disabled")
        radio_button_10.configure(state="disabled")
        radio_button_11.configure(state="disabled")
        updateuseai(True)
    rememberload_var = tk.IntVar(value=int(value[14][:-1]))
    check_2 = CTkCheckBox(settingstab.tab(lang[166]), text=lang[40], variable=rememberload_var, onvalue=1, offvalue=0, font=CTkFont(size=15),command=rememberload) #Remember Last Project
    compact_var = tk.IntVar(value=compact)
    check_4 = CTkCheckBox(settingstab.tab(lang[165]), text=lang[41], variable=compact_var, onvalue=1, offvalue=0, font=CTkFont(size=15),command=updatesidebar) #Collapse Side Bar
    overexport_var = tk.IntVar(value=overexport)
    check_5 = CTkCheckBox(settingstab.tab(lang[166]), text=lang[43], variable=overexport_var, onvalue=1, offvalue=0, font=CTkFont(size=15),command=overexporter) #Overwrite Old Exports
    askexport_var = tk.IntVar(value=askexport)
    check_6 = CTkCheckBox(settingstab.tab(lang[166]), text=lang[44], variable=askexport_var, onvalue=1, offvalue=0, font=CTkFont(size=15),command=askexporter) #Ask for File Names before Exporting
    unrestrained_var = tk.IntVar(value=use_manual)
    check_7 = CTkCheckBox(settingstab.tab(lang[167]), text=lang[45], variable=unrestrained_var, onvalue=1, offvalue=0, font=CTkFont(size=15),command=usemanual) #Manual Mode
    welcome_var = tk.IntVar(value=int(value[10][:-1]))
    check_8 = CTkCheckBox(settingstab.tab(lang[165]), text=lang[53], variable=welcome_var, onvalue=0, offvalue=1, font=CTkFont(size=15),command=updatewelcome) #Welcome Screen

    #Export Elements
    text53 = CTkLabel(exportframe, text=lang[46], font=CTkFont(size=15)) #Step 1
    text54 = CTkLabel(exportframe, text=lang[47], font=CTkFont(size=12))
    exportsave = CTkButton(exportframe,text=lang[55],font=CTkFont(size=bsize),width=bwidth,height=bheight,command=exportsaver) #Export Folder
    export1 = CTkButton(expframe1,text=lang[58],font=CTkFont(size=bsize),width=bwidth,height=bheight,command=exportxlsx,state='disabled') #Export CSV
    export2 = CTkButton(expframe1,text=lang[56],font=CTkFont(size=bsize),width=bwidth,height=bheight,command=exportcsv,state='disabled') #Export XMFV
    export5 = CTkButton(expframe1,text=lang[57],font=CTkFont(size=bsize),width=bwidth,height=bheight,command=exportxmfv,state='disabled') #Export XLSX
    exportsuccess = CTkLabel(expframe,text='',image=imgtk18)
    exportfail = CTkLabel(expframe,text='',image=imgtk19)
    exportclick = CTkLabel(expframe,text='',image=imgtk20)
    export3 = CTkButton(expframe,text=lang[59],font=CTkFont(size=bsize),width=bwidth,height=bheight,command=openexplorer,state='normal') #Show in Explorer
    export4 = CTkButton(expframe,text=lang[60],font=CTkFont(size=bsize),width=bwidth,height=bheight,command=dynafram,state='normal') #Open DynaFRAM

    #Analysis Elements
    text12 = CTkLabel(analyzeframe, text=lang[61], font=CTkFont(size=15)) #Step 1: Load SD Card Folder
    text13 = CTkLabel(analyzeframe, text=lang[62], font=CTkFont(size=12)) #Load your desired folder
    text42 = CTkLabel(analyzeframe, text=lang[63], font=CTkFont(size=15)) #Step 2: Start Analysis
    text43 = CTkLabel(analyzeframe, text=lang[64], font=CTkFont(size=12)) #Click Analyze Now
    text84 = CTkLabel(progressframe, text=lang[164], font=CTkFont(size=15)) #Progress
    text85 = CTkLabel(progressframe, text=lang[63], font=CTkFont(size=12)) #Desc
    text86 = CTkLabel(progressframe, text="0%", font=CTkFont(size=12)) #%
    progressbar = CTkProgressBar(progressframe)
    progressbar.set(0)
    text44 = CTkLabel(imgframe, text="\n"*46, font=CTkFont(size=15))
    loadfile = CTkButton(analyzeframe,text=lang[65],font=CTkFont(size=bsize),width=bwidth,height=bheight,command=loadjson) #Load Folder
    analyze = CTkButton(analyzeframe,text=lang[66],font=CTkFont(size=bsize),width=bwidth,height=bheight,state="disabled",command=start_analysis) #Analyze Now
    text45 = CTkLabel(imgframe2, text="", font=CTkFont(size=15))
    text50 = CTkLabel(imgframe4, text=lang[67], font=CTkFont(size=15)) #Head Orientation Times
    headup = CTkLabel(imgframe4,text=lang[68],image=imgtk6,compound='left',font=CTkFont(size=15),justify='left',padx=10) #Head Up Time
    headlevel = CTkLabel(imgframe4,text=lang[69],image=imgtk30,compound='left',font=CTkFont(size=15),justify='left',padx=10) #Head Level Time
    headdown = CTkLabel(imgframe4,text=lang[70],image=imgtk31,compound='left',font=CTkFont(size=15),justify='left',padx=10) #Head Down Time
    editnames = CTkButton(imgframe6,text=lang[71],font=CTkFont(size=bsize),width=bwidth,height=bheight,command=editAOI) #Edit Names
    text51 = CTkLabel(imgframe5, text=lang[72]+": 0", font=CTkFont(size=15)) #No-Go Zone Violations
    showgraph = CTkButton(imgframe3,text=lang[74],font=CTkFont(size=bsize),width=bwidth,height=bheight,command=showgraphs) #Show Graphs
    violateradar = CTkButton(imgframe3,text=lang[75],font=CTkFont(size=bsize),width=bwidth,height=bheight,command=radar) #Radar Violations
    deadmanbutton = CTkButton(imgframe3,text=lang[51],font=CTkFont(size=bsize),width=bwidth,height=bheight) #Dead Man Switch
    detectionsbutton = CTkButton(imgframe6,text=lang[52],font=CTkFont(size=bsize),width=bwidth,height=bheight,command=detectionsfun) #Detections
    clearbutton = CTkButton(imgframe6,text=lang[50],font=CTkFont(size=bsize),width=bwidth,height=bheight,fg_color='dark red',hover_color='#530000',command=clearbuttonfun) #Clear All

    #Record Elements
    preview_var = tk.IntVar(value=int(value[7][:-1]))
    switch_1 = CTkSwitch(recordframe, text=lang[26], variable=preview_var, onvalue=1, offvalue=0, font=CTkFont(size=15),command=previewtoggle) #Pop Live Preview
    connectbutton = CTkButton(recordframe,text=lang[78],command=lambda : asyncio.run(connect()),font=CTkFont(size=bsize),width=187,height=bheight)
    calibratebutton = CTkButton(recordframe,text=lang[81],command=lambda : asyncio.run(calibrate()),state="disabled",font=CTkFont(size=bsize),width=187,height=bheight)
    recordbutton = CTkButton(recordframe,text=lang[84],state="disabled",font=CTkFont(size=bsize),command=lambda: asyncio.run(record()),width=187,height=bheight,fg_color='dark red',hover_color='#530000') #Record
    text20 = CTkLabel(videoframe, text=lang[85], font=CTkFont(size=15)) #Live View
    videolabel = Label(videoframe)
    previewpop = CTkLabel(videoframe,text='')
    connectpreview = CTkLabel(videoframe,text='')
    previewavail = CTkLabel(videoframe,text='')
    previewload = ImageLabel(videoframe)
    recordstatus = CTkLabel(videoframe,text='')
    text21 = CTkLabel(videoframe, text=lang[86], font=CTkFont(size=12))
    text22 = CTkLabel(videoframe, text=lang[86], font=CTkFont(size=12))
    batteryicon = CTkLabel(videoframe,text='')
    timelefticon = CTkLabel(videoframe,text='')
    previewbutton = CTkButton(recordframe,text=lang[88],command=video,font=CTkFont(size=bsize),width=187,height=bheight, state="disabled") #Start Live View

    framebuttons = [settingsbutton,recorddata,exportdata,analyzedata]

    #Main Interface
    frame1.pack(fill=tk.Y,side=tk.LEFT)
    if compact==0:
        text80.pack(side=tk.TOP,pady=(20,0))
        heading1.pack(side=tk.TOP,pady=(10,0))
        heading2.pack(side=tk.TOP,pady=(0,10))
        recorddata.pack(side=tk.TOP,pady=10,padx=20)
        analyzedata.pack(side=tk.TOP,pady=10,padx=20)
        viewdata.pack(side=tk.TOP,pady=10,padx=20)
        exportdata.pack(side=tk.TOP,pady=10,padx=20)
        settingsbutton.pack(side=tk.TOP,pady=10,padx=20)
        text1.pack(side=tk.BOTTOM,pady=10,padx=20)
        quitbutton.pack(side=tk.BOTTOM,pady=(10,0),padx=20)
    else:
        recorddata.configure(text="",height=bsheight,width=bswidth,corner_radius=0,compound="bottom")
        recorddata.pack(side=tk.TOP,pady=0,padx=0)
        analyzedata.configure(text="",height=bsheight,width=bswidth,corner_radius=0,compound="bottom")
        analyzedata.pack(side=tk.TOP,pady=0,padx=0)
        viewdata.configure(text="",height=bsheight,width=bswidth,corner_radius=0,compound="bottom",fg_color="#333333")
        viewdata.pack(side=tk.TOP,pady=0,padx=0)
        exportdata.configure(text="",height=bsheight,width=bswidth,corner_radius=0,compound="bottom")
        exportdata.pack(side=tk.TOP,pady=0,padx=0)
        settingsbutton.configure(text="",height=bsheight,width=bswidth,corner_radius=0,compound="bottom")
        settingsbutton.pack(side=tk.TOP,pady=0,padx=0)
        quitbutton.configure(text="",height=bsheight,width=bswidth,corner_radius=0,compound="bottom")
        quitbutton.pack(side=tk.BOTTOM,pady=(0,0),padx=0)
        usermanual2.configure(text="",height=bsheight,width=bswidth,corner_radius=0,compound="bottom")
    
    autoloadfiles()
    if welcome:
        welcomepage()
    else:
        recordpage()
    root.after(500,fullscreen)

    root.mainloop()

if __name__=='__main__':
    main()