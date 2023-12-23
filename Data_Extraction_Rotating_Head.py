#Data Extraction Rotating Head
#Code part of Tobii Eye Tracking Project
#Made by Akash Samanta

import gzip
import json
import openpyxl as xls
import os
import math

def timestamper(string):
    dict1 = json.loads(string)
    return dict1["timestamp"]

def gygp3ac(gp3,gy,sdir):
    Timestamp,Gyro_X,Gyro_Y,Gyro_Z,X,Y,Z,X2,Y2=[],[],[],[],[],[],[],[],[]
    wb = xls.Workbook()
    ws = wb.active
    ws.title = "Data"
    c1 = ws.cell(row=1, column=1)
    c1.value = "Computer Timestamp [ms]"
    c2 = ws.cell(row=1, column=2)
    c2.value = "Gaze point 3D X [HUCS mm]"
    c3 = ws.cell(row=1, column=3)
    c3.value = "Gaze point 3D Y [HUCS mm]"
    c4 = ws.cell(row=1, column=4)
    c4.value = "Gaze point 3D Z [HUCS mm]"
    c5 = ws.cell(row=1, column=5)
    c5.value = "Gyro X [°/s]"
    c6 = ws.cell(row=1, column=6)
    c6.value = "Gyro Y [°/s]"
    c7 = ws.cell(row=1, column=7)
    c7.value = "Gyro Z [°/s]"
    c8 = ws.cell(row=1, column=8)
    c8.value = "Gaze Point 2D X [Pixels]"
    c9 = ws.cell(row=1, column=9)
    c9.value = "Gaze Point 2D Y [Pixels]"

    tmin=[]
    tmin.append(timestamper(gp3[0]))
    tmin.append(timestamper(gy[0]))
    tmin=min(tmin)

    tmax=[]
    tmax.append(timestamper(gp3[-1]))
    tmax.append(timestamper(gy[-1]))
    tmax=max(tmax)

    index1,index2=0,0
    while True:
        try:     
            if index1>len(gp3)-1:
                gp3x,gp3y,gp3z,gpx,gpy=0.00,0.00,0.00,0.00,0.00
            else:
                if timestamper(gp3[index1])<=timestamper(gy[index2]):
                    data=gp3[index1]
                    data=json.loads(data)
                    data=dict(data["data"])
                    try:
                        gp3x=float(data["gaze3d"][0])
                        gp3y=float(data["gaze3d"][1])
                        gp3z=float(data["gaze3d"][2])
                        gpx=float(data["gaze2d"][0])
                        gpy=float(data["gaze2d"][1])
                    except:
                        index1=index1+1
                        continue
                    index1=index1+1
                else:
                    gp3x,gp3y,gp3z,gpx,gpy=0.00,0.00,0.00,0.00,0.00
            
            if index2>len(gy)-1:
                gyx,gyy,gyz=0.00,0.00,0.00
            else:
                if timestamper(gy[index2])<timestamper(gp3[index1]):
                    data=gy[index2]
                    data=json.loads(data)
                    data=dict(data["data"])
                    try:
                        gyx=float(data["gyroscope"][0])
                        gyy=float(data["gyroscope"][1])
                        gyz=float(data["gyroscope"][2])
                    except:
                        index2=index2+1
                        continue
                    index2=index2+1
                else:
                    gyx,gyy,gyz=0.00,0.00,0.00

            if gyx==0.00 or gyy==0.00 or gyz==0.00 or gp3x==0.00 or gp3y==0.00 or gp3z==0.00 or gpx==0.00 or gpy==0.00:
                pass
            else:
                Timestamp.append(min(timestamper(gy[index2]),timestamper(gp3[index1]))*1000000)
                Gyro_X.append(gyx)
                Gyro_Y.append(gyy)
                Gyro_Z.append(gyz)
                X.append(gp3x)
                Y.append(gp3y)
                Z.append(gp3z)
                X2.append(gpx)
                Y2.append(gpy)
        except IndexError:
            break

    for i in range(len(Timestamp)-1):
        cell = ws.cell(row=i+2,column=1)
        cell.value = Timestamp[i]
        cell = ws.cell(row=i+2,column=2)
        cell.value = X[i]
        cell = ws.cell(row=i+2,column=3)
        cell.value = Y[i]
        cell = ws.cell(row=i+2,column=4)
        cell.value = Z[i]
        cell = ws.cell(row=i+2,column=5)
        cell.value = Gyro_X[i]
        cell = ws.cell(row=i+2,column=6)
        cell.value = Gyro_Y[i]
        cell = ws.cell(row=i+2,column=7)
        cell.value = Gyro_Z[i]
        cell = ws.cell(row=i+2,column=8)
        cell.value = X2[i]
        cell = ws.cell(row=i+2,column=9)
        cell.value = Y2[i]

    save=sdir+'/Glass Data Export.xlsx'
    wb.save(save)
    return save

def extract(folder):
    if os.path.exists(folder+'/Glass Data Export.xlsx'):
        save=folder+'/Glass Data Export.xlsx'
        return save
    else:
        f=gzip.open(folder+"\gazedata.gz", "r")
        imu_data,gaze_data = [],[]
        for line in f:
            text=line.decode('UTF-8')
            gaze_data.append(text)
        f.close()
        f=gzip.open(folder+"\imudata.gz", "r")
        for line in f:
            text=line.decode('UTF-8')
            imu_data.append(text)
        f.close()
        return gygp3ac(gaze_data,imu_data,folder)

if __name__=='__main__':
    print("Data Extraction Rotating Head\nCode part of Tobii Eye Tracking Project\nMade by Akash Samanta\n")
    print("WARNING! You are running this code snippet alone.\nThis may accomplish only a specific task. Please use EyeFRAM GUI.py instead.")
    extract(r'C:\Users\akash\Documents\Work Files\EyePort New\First Recording',1)
