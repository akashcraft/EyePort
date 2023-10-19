#Data Collection Rotating Head
#Code part of Tobii Eye Tracking Project
#Made by Akash Samanta

import gzip
import json
import openpyxl as xls
import os
import math

def timestamper(string):
    ans=''
    for j in string[7:]:
        if j==',':
            break
        else:
            ans=ans+j
    return ans

def gygp3(gy,gp3,gp,sdir,overwrite):
    Timestamp,Gyro_X,Gyro_Y,Gyro_Z,X,Y,Z,X2,Y2=[],[],[],[],[],[],[],[],[]
    wb = xls.Workbook()
    ws = wb.active
    ws.title = "Data"
    c1 = ws.cell(row=1, column=1)
    c1.value = "Computer Timestamp"
    c2 = ws.cell(row=1, column=2)
    c2.value = "Gaze Point 3D X [HUCS mm]"
    c3 = ws.cell(row=1, column=3)
    c3.value = "Gaze Point 3D Y [HUCS mm]"
    c4 = ws.cell(row=1, column=4)
    c4.value = "Gaze Point 3D Z [HUCS mm]"
    c5 = ws.cell(row=1, column=5)
    c5.value = "Gyro X [°/s]"
    c6 = ws.cell(row=1, column=6)
    c6.value = "Gyro Y [°/s]"
    c7 = ws.cell(row=1, column=7)
    c7.value = "Gyro Z [°/s]"
    c8 = ws.cell(row=1, column=8)
    c8.value = "Gaze Point 2D X [Pixels]"
    c9 = ws.cell(row=1, column=9)
    c9.value = "Gaze Point 2D Y [Pixels]"

    tgy,tgp3,tgp=[],[],[]
    for i in gy:
        if i not in tgy:
            tgy.append(i)
    for i in gp3:
        if i not in tgp3:
            tgp3.append(i)
    for i in gp:
        if i not in tgp:
            tgp.append(i)
    gy=tgy
    gp3=tgp3
    gp=tgp

    tmin=[]
    if len(gp3)!=0:
        tmin.append(round(float(timestamper(gp3[0]))/1000))
        tmin.append(round(float(timestamper(gp[0]))/1000))
    tmin.append(round(float(timestamper(gy[0]))/1000))
    tmin=min(tmin)

    tmax=[]
    if len(gp3)!=0:
        tmax.append(round(float(timestamper(gp3[-1]))/1000))
        tmax.append(round(float(timestamper(gp[-1]))/1000))
    tmax.append(round(float(timestamper(gy[-1]))/1000))
    tmax=max(tmax)

    index1,index2,index4=0,0,0
    for i in range(tmax-tmin+1):
        if index1>len(gp3)-1:
            gp3x,gp3y,gp3z=0.00,0.00,0.00
            gpx,gpy=0.00,0.00
        else:
            try:
                if round(float(timestamper(gp3[index1]))/1000)==(tmin+i):
                    data=gp3[index1]
                    data=data[1:-2]
                    data=data.split(':')
                    data=data[-1][2:-1].split(',')
                    gp3x=float(data[0])
                    gp3y=float(data[1])
                    gp3z=float(data[2])
                    data=gp[index1]
                    data=data[1:-2]
                    data=data.split(':')
                    data=data[-1][2:-1].split(',')
                    gpx=float(data[0])
                    gpy=float(data[1])
                    index1=index1+1
                else:
                    gp3x,gp3y,gp3z=0.00,0.00,0.00
                    gpx,gpy=0.00,0.00
            except:
                gp3x,gp3y,gp3z=0.00,0.00,0.00
                gpx,gpy=0.00,0.00
        
        if index2>len(gy)-1:
            gyx,gyy,gyz=0.00,0.00,0.00
        else:
            try:
                if round(float(timestamper(gy[index2]))/1000)==(tmin+i):
                    data=gy[index2]
                    data=data[1:-2]
                    data=data.split(':')
                    data=data[-1][2:-1].split(',')
                    gyx=float(data[0])
                    gyy=float(data[1])
                    gyz=float(data[2])
                    index2=index2+1
                else:
                    gyx,gyy,gyz=0.00,0.00,0.00
            except:
                gyx,gyy,gyz=0.00,0.00,0.00

        if gyx==0.00 and gyy==0.00 and gyz==0.00 and gp3x==0.00 and gp3y==0.00 and gp3z==0.00 and gpx==0.00 and gpy==0.00:
            pass
        else:
            Timestamp.append(tmin+i)
            Gyro_X.append(gyx)
            Gyro_Y.append(gyy)
            Gyro_Z.append(gyz)
            X.append(gp3x)
            Y.append(gp3y)
            Z.append(gp3z)
            X2.append(gpx)
            Y2.append(gpy)
      
    for i in range(len(Timestamp)):
        cell = ws.cell(row=i+2,column=1)
        cell.value = Timestamp[i]
        cell = ws.cell(row=i+2,column=2)
        cell.value = X[i]
        cell = ws.cell(row=i+2,column=3)
        cell.value = Y[i]
        cell = ws.cell(row=i+2,column=4)
        cell.value = Z[i]
        cell = ws.cell(row=i+2,column=5)
        cell.value = Gyro_X[i]
        cell = ws.cell(row=i+2,column=6)
        cell.value = Gyro_Y[i]
        cell = ws.cell(row=i+2,column=7)
        cell.value = Gyro_Z[i]
        cell = ws.cell(row=i+2,column=8)
        cell.value = X2[i]
        cell = ws.cell(row=i+2,column=9)
        cell.value = Y2[i]
    c=1
    while True:
        if os.path.exists(sdir+'/'+f'Project{c} Data Export.xlsx'):
            c=c+1
        else:
            break
    if c==1:
        save=sdir+'/'+f'Project1 Data Export.xlsx' #c-1 to OVERWRITE c to make NEW
    else:
        if overwrite==1:
            save=sdir+'/'+f'Project{c-1} Data Export.xlsx' #c-1 to OVERWRITE c to make NEW
        else:
            save=sdir+'/'+f'Project{c} Data Export.xlsx'
    wb.save(save)

def collect(recsavefolder,data,overwrite):
    gp3_data,gy_data,gp_data = [],[],[]
    for text in data:
        if "'gp3':" in str(text):
            gp3_data.append(str(text))
        if "'gy':" in str(text):
            gy_data.append(str(text))
        if "'gp':" in str(text):
            gp_data.append(str(text))
    #print(len(gp3_data),len(gy_data),len(gp_data))
    gygp3(gy_data,gp3_data,gp_data,recsavefolder,overwrite)

if __name__=='__main__':
    print("Data Collection Rotating Head\nCode part of Tobii Eye Tracking Project\nMade by Akash Samanta\n")
    print("WARNING! You are running this code snippet alone.\nThis may accomplish only a specific task. Please use EyeFRAM GUI.py instead.")

