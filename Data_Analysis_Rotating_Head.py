#Data Analysis Rotating Head
#Stable - Can also receive Data from Unrestrained Cases but may not give desired results
#Code part of Tobii Eye Tracking Project
#Made by Akash Samanta

from image_similarity_measures.evaluate import evaluation
from customtkinter import *
import openpyxl as xls
import numpy as np
import os,shutil,json,cv2
from threading import Thread
from queue import Queue
imageai_avail = True
try:
    from CustomObjectDetector import CustomObjectDetect
    from ObjectDetector import ObjectDetect
except:
    imageai_avail = False
from PIL import Image

progress = 0
progress_queue = Queue()
__all__ = ['progress_queue']

#Functions
def headtable(up,level,down):
    global headduration
    uptime = 0
    for i in up:
        uptime = uptime + (i[1]-i[0])
    leveltime = 0
    for i in level:
        leveltime = leveltime + (i[1]-i[0])
    downtime = 0
    for i in down:
        downtime = downtime + (i[1]-i[0])

    headduration = [round(uptime,2),round(leveltime,2),round(downtime,2)]

    for i in up:
        i.insert(0,"Head Up")
    for i in level:
        i.insert(0,"Head Level")
    for i in down:
        i.insert(0,"Head Down")
    all = []
    for i in [up,level,down]:
        for j in i:
            all.append(j)
    
    for i in range(len(all)-1):
        for j in range(len(all)-1-i):
            if all[j][1]>all[j+1][1]:
                temp = all[j+1]
                all[j+1] = all[j]
                all[j] = temp
                
    return all

def unique_analyzer_diesel(answer):
    if len(answer)==0:
        return 0,[0,],[],[]
    else:
        UAOIName = []
        for i in answer:
            if i not in UAOIName:
                UAOIName.append(i)
        pos = [0,]
        prev=answer[0]
        for i in range(len(answer)):
            if answer[i]!=prev:
                pos.append(i)
                prev = answer[i]
        return len(UAOIName),pos,answer,UAOIName

def unique_analyzer(L,sen):
    if len(L)==0:
        return 0,[0,],[],[]
    elif len(L)==1:
        return 1,[0,],['Object 1'],['Object 1']
    else:
        Result=[L[0],]
        pos=[0,] #Main index to use when showing unique areas of interest
        AOIName=['Object 1',]
        UAOIName=['Object 1',]
        ucount=2
        for i in range(1,len(L)):
            for j in range(len(Result)):
                error = evaluation(L[i],Result[j],["ssim"])
                if error['ssim']>sen: #Same Object
                    flag=0
                    AOIName.append('Object '+str(j+1))
                    break
                else:
                    flag=1 #Different Object
            if flag==1:
                Result.append(L[i])
                AOIName.append(f'Object {ucount}')
                UAOIName.append(f'Object {ucount}')
                ucount=ucount+1
                pos.append(i)
                flag=0

        #print(pos) 
        return (len(Result)),pos,AOIName,UAOIName

def analyzer(sdir,vdir,sen=0.75,useai=0,manual=0,thit=200,headmode=2,headsen=20,atolbox=600): #Data Directory, Video Directory, Sensitivity for UAOI, Fixation Time, Use AI?, Type of Recording, Head Orientation
    global progress
    #Counters and AOIs
    GX=[] #All Gyro_X D/s
    GY=[] #All Gyro_Y D/s
    GZ=[] #All Gyro_Z D/s
    X,Y,Z=[],[],[] #All 3D Gaze
    X2,Y2=[],[] #All 2D Gaze
    Timestamps=[] #All Timestamps

    if useai!=0 and not imageai_avail:
        useai=0

    #Check if Gaze Overlays are needed
    current_dir = os.path.dirname(sdir)
    g3_file = current_dir+"/recording.g3"
    if (os.path.exists(g3_file)):
        if not os.path.exists(current_dir+"/recording.txt"):
            shutil.copyfile(g3_file,current_dir+"/recording.txt")
        fobj = open(current_dir+"/recording.txt",'r')
        gaze_enabled = json.loads(fobj.read())['scenecamera']['gaze-overlay']
        fobj.close()
    else:
        raise FileNotFoundError
    
    progress = 5
    progress_queue.put(progress)
    
    #Column Settings
    columnselect=[1,2,3,4,5,6,7,8,9]

    #Loading Excel File
    book = xls.load_workbook(sdir)
    #Getting Sheet
    sheet = book['Data']
    col = sheet.iter_cols

    #Data Collection begins
    #sheet.max_row:
    for j in columnselect:
        for col in sheet.iter_cols(j,j):
            for i in range(1, sheet.max_row):
                if j==2:
                    X.append(col[i].value)
                elif j==3:
                    Y.append(col[i].value)
                elif j==4:
                    Z.append(col[i].value)
                elif j==5:
                    GX.append(col[i].value)
                elif j==6:
                    GY.append(col[i].value)
                elif j==7:
                    GZ.append(col[i].value)
                elif j==8:
                    X2.append(col[i].value)
                elif j==9:
                    Y2.append(col[i].value)
                else:
                    Timestamps.append(col[i].value)

    progress = 10
    progress_queue.put(progress)
        
    #Calculating Head Up, Head Level, Head Down Times
    headup,headlevel,headdown = [],[[0,],],[]
    state = headmode
    if state==2: #From Level Start
        headup,headlevel,headdown = [],[[0,],],[]
    elif state==3: #From Down Start
        headup,headlevel,headdown = [],[],[[0,],]
    else: #From Up Start
        headup,headlevel,headdown = [[0,],],[],[]
    ready = True
    starter = False
    #headsen User defined acceleration value
    for i in range(len(GX)):
        if GX[i]<-headsen: #Head Motion Up
            if ready:
                if state==2: #From Level
                    headlevel[-1].append(Timestamps[i]/1000000)
                    state=1
                    starter=True
                elif state==3: #From Down
                    headdown[-1].append(Timestamps[i]/1000000)
                    state=2
                    starter=True
                else:
                    state=1
                ready = False
        elif GX[i]>headsen: #Head Motion Down
            if ready:
                if state==1: #From Up
                    headup[-1].append(Timestamps[i]/1000000)
                    state=2
                    starter=True
                elif state==2: #From Level
                    headlevel[-1].append(Timestamps[i]/1000000)
                    state=3
                    starter=True
                else:
                    state=3
                ready = False
        else: #Head Level in one of the three states
            if starter:
                starter=False
                if state==1:
                    headup.append([Timestamps[i]/1000000,])
                elif state==2:
                    headlevel.append([Timestamps[i]/1000000,])
                else:
                    headdown.append([Timestamps[i]/1000000,])
            ready=True
   
    #Add missing last timestamps
    for i in [headup,headlevel,headdown]:
        if len(i)!=0 and len(i[-1])!=2:
            i[-1].append(Timestamps[-1]/1000000)

    progress = 25
    progress_queue.put(progress)

    #Analysis Begins
    #Main Algorithm - AOI Detections
    tripper=0 #Trip counter to capture AOI. This value should match with thit - the fixation duration. Not returned.
    count=0 #Counts total number of AOI. This is returned.
    if useai==3:
        s=0.015 #Sensitivity in pixels (Worked many times) User cannot change this setting. But you as a developer can!
    else:
        s=0.025
    #thit = 200 - Around 2 Second default. User can change this setting.
    GazeSecond=[]
    End=[]
    GazeX,GazeY=[],[]
    for i in range(1,len(Timestamps)):
        if tripper==thit:
            #FOUND A NEW AOI
            GazeSecond.append(Timestamps[i])
            GazeX.append(X2[i])
            GazeY.append(Y2[i])
            count=count+1
        
        X2upper = X2[i-1]+s
        X2lower = X2[i-1]-s
        Y2upper = Y2[i-1]+s
        Y2lower = Y2[i-1]-s
        if X2[i]<X2upper and X2[i]>X2lower and Y2[i]<Y2upper and Y2[i]>Y2lower:
            tripper=tripper+1
        else:
            if tripper>=thit:
                End.append(Timestamps[i]) #END OF AOI
            tripper=0
    
    #In case the last object does not have a End Time Stamp
    if len(GazeSecond)-len(End)==1:
        End.append(Timestamps[-1])

    progress = 40
    progress_queue.put(progress)

    startsec=Timestamps[0] #Starting Timestamp
    picarray=[] #For Frames with Gaze
    nopicarray=[] #For Frames without Gaze NOT RETURNED but needed for ObjectDetector.py
    answer=[] #To store all object detection text answers
    answerarray=[] #To store object detection frames

    #Video File Handles
    video = cv2.VideoCapture(vdir)
    fps = video.get(cv2.CAP_PROP_FPS)
    videoduration = video.get(cv2.CAP_PROP_FRAME_COUNT)
    videoduration = videoduration/fps

    #Syncronization Step - To align video GazeSecond and timestamps
    ratio=videoduration/((Timestamps[-1]-startsec)/1000000)
    for i in range(len(GazeSecond)):
        GazeSecond[i]=((GazeSecond[i]-startsec)/1000000)*ratio
    for i in range(len(End)):
        End[i]=((End[i]-startsec)/1000000)*ratio

    progress = 50
    progress_queue.put(progress)

    #Getting Frames and Adding Overlays
    for i in range(len(GazeSecond)): #Get all frames
        if GazeSecond[0]<0.1: #Fix the Green Screen Issue at 0.0s
            video.set(cv2.CAP_PROP_POS_FRAMES,int(fps*0.2))
        else:
            video.set(cv2.CAP_PROP_POS_FRAMES,int(fps*GazeSecond[i]))
        ret,frame=video.read()
        if not gaze_enabled: #If recorded in without overlays, then add the gaze circles 
            x=GazeX[i]
            y=GazeY[i]
            if x>1:
                x=1
            if y>1:
                y=1
            frame=cv2.circle(frame,(int(x*1920),int(y*1080)), 25, (0, 0, 255), 3)
        try:
            frame=cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        except:
            video.set(cv2.CAP_PROP_POS_FRAMES,int((fps*GazeSecond[i])-30))
            ret,frame=video.read()
            try: #Last Frame Video Bug, Go Back 30 fps or 60 fps and try again
                frame=cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            except:
                video.set(cv2.CAP_PROP_POS_FRAMES,int((fps*GazeSecond[i])-60))
                ret,frame=video.read()
                frame=cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

        picarray.append(frame)

    progress = 60
    progress_queue.put(progress)

    #Unique AOI Detection
    cropped = []
    if not os.path.exists(current_dir+"/detections"):
            os.makedirs(current_dir+"/detections")
    else:
        delfiles = os.listdir(current_dir+"/detections")
        for j in range(len(delfiles)):
            delfiles[j] = os.path.join(current_dir+"/detections",delfiles[j])
        for j in delfiles:
            os.remove(j)
    for i in range(len(GazeSecond)):
        s = atolbox
        x1 = int(int(GazeX[i]*1920)-int(s/2))
        y1 = int(int(GazeY[i]*1080)-int(s/2))
        x2 = x1 + s
        y2 = y1 + s
        if x1<0: #Check if slice is going out of bounds
            x1,x2=0,s
        if y1<0:
            y1,y2=0,s
        if x2>1920:
            x2,x1=1920,int(1920-s)
        if y2>1080:
            y2,y1=1080,int(1080-s)
        image = picarray[i]
        image = cv2.cvtColor(image,cv2.COLOR_BGR2RGB)
        dest = current_dir+f"/detections/Object {i+1}.jpeg"
        cv2.imwrite(dest, image[y1:y2, x1:x2])
        cropped.append(dest)

    progress = 70
    progress_queue.put(progress)

    #No Go Zones
    i=1
    violation=[]
    current_dir = os.getcwd()+r"/NoZones/"
    if not os.path.exists(current_dir):
        os.mkdir(current_dir)
    check = os.listdir(current_dir)
    if len(check)!=0:
        for i in range(len(check)):
            check[i] = os.path.join(current_dir,check[i])
        flag=False
        delflag=False
        for i in check:
            for j in cropped:
                try:
                    error = evaluation(i,j,["ssim"])
                except:
                    delflag = True
                    break
                if error['ssim']>0.90:
                    flag=True
                    break
            if delflag:
                break
            if flag:
                flag=False
                violation.append(i)
        if delflag:
            for i in check:
                os.remove(i)

    progress = 75
    progress_queue.put(progress)
        
    if useai==1: #If AI needed, then call ObjectDetector.py
        for i in range(len(GazeSecond)):
            video.set(cv2.CAP_PROP_POS_FRAMES,int(fps*GazeSecond[i]))
            ret,frame=video.read()
            frame=cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            nopicarray.append(frame)
            text,returned_image = ObjectDetect(frame, GazeX[i], GazeY[i])
            answer.append(text.capitalize())
            if not gaze_enabled: #If recorded in without overlays, then add the gaze circles
                x=GazeX[i]
                y=GazeY[i]
                if x>1:
                    x=1
                if y>1:
                    y=1
                returned_image=cv2.circle(returned_image,(int(x*1920),int(y*1080)), 40, (255, 0, 0), 3)
            answerarray.append(returned_image)
    elif useai==2 or useai==3: #If AI needed, then call CustomObjectDetector.py for Ships and Icebergs / VISTA Diesel Engine
        for i in range(len(GazeSecond)):
            video.set(cv2.CAP_PROP_POS_FRAMES,int(fps*GazeSecond[i]))
            ret,frame=video.read()
            if useai==2:
                frame=cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            nopicarray.append(frame)
            text,returned_image = CustomObjectDetect(frame, GazeX[i], GazeY[i],useai)
            if useai==2:
                answer.append(text.capitalize())
            else:
                answer.append(text)
            if not gaze_enabled: #If recorded in without overlays, then add the gaze circles
                x=GazeX[i]
                y=GazeY[i]
                if x>1:
                    x=1
                if y>1:
                    y=1
                returned_image=cv2.circle(returned_image,(int(x*1920),int(y*1080)), 40, (255, 0, 0), 3)
            answerarray.append(returned_image)
    else:
        pass

    progress = 90
    progress_queue.put(progress)

    #Send all frames
    if manual==0:
        if useai==3:
            unique_count,pos,AOIName,UAOIName = unique_analyzer_diesel(answer)
        else:
            unique_count,pos,AOIName,UAOIName = unique_analyzer(cropped,sen)
    else:
        unique_count,pos,AOIName,UAOIName=count,[i for i in range(count)],[f'Object {i+1}' for i in range(count)],[f'Object {i+1}' for i in range(count)]

    #Information for Gaze 2D Graph Color and Size
    tX,tY,color,size=[],[],[],[]
    for i in range(len(X2)):
        if X2[i]!=0:
            if X2[i] in GazeX and Y2[i] in GazeY:
                color.append(2)
                size.append(1000)
            else:
                color.append(1)
                size.append(100)
            tX.append(X2[i])
            tY.append(Y2[i])

    progress = 95
    progress_queue.put(progress)

    X2,Y2=tX,tY
    #print(answer,unique_count,AOIName,UAOIName)
    return count,unique_count,picarray,answerarray,answer,pos,AOIName,UAOIName,GazeSecond,End,X2,Y2,Timestamps,color,size,headtable(headup,headlevel,headdown),headduration,violation,imageai_avail

if __name__=='__main__':
    print("Data Analyzer Rotating Head\nCode part of Tobii Eye Tracking Project\nMade by Akash Samanta\n")
    print("WARNING! You are running this code snippet alone.\nThis may accomplish only a specific task. Please use EyeFRAM GUI.py instead.\n")
    #c=1
    #count,unique_count,picarray,answerarray,answer,pos,AOIName,UAOIName,GazeSecond,End,X,Y,color,size=analyzer(
        #sdir=rf'D:\20230925T142216Z\Project{c} Data Export.xlsx',
        #vdir=rf'D:\20230925T142216Z\scenevideo.mp4',
        #sen=0.5,
        #useai=0,
        #manual=0)
    #print("The System has detected",count,"area of interests in the given session\nWith",unique_count,"of them being unique.\n")


